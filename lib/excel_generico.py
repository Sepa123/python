from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font , PatternFill, Border ,Side
from openpyxl.worksheet.page import PageMargins

from routers.logistica_inversa import cambiar_bool 

# def cambiar_bool(mi_tupla):
#     lista= []

#     for valor in mi_tupla:

#         if isinstance(valor, bool):
#             if valor is True:
#                 lista.append('x')
#             else:
#                 lista.append('')
#         else:
#             lista.append(valor)

#     return tuple(lista)
        

def generar_excel_generico(results, nombre_filas, nombre_excel ):

    # results = conn.obtener_etiqueta_carga_rsv(nombre_carga,codigo,tipo)
    wb = Workbook()
    ws = wb.active
    results.insert(0, nombre_filas)

    for row in results:
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(f"excel/{nombre_excel}.xlsx")

    return FileResponse(f"excel/{nombre_excel}.xlsx")

def generar_excel_con_titulo(results, nombre_filas, nombre_excel,titulo ):

    # results = conn.obtener_etiqueta_carga_rsv(nombre_carga,codigo,tipo)
    wb = Workbook()
    ws = wb.active

    margins = PageMargins(top=0.3, bottom=0.6, left=0.4, right=0.5, header=0.3, footer=0.3)
    ws.page_margins = margins

    # Estilo para el texto en negrita
    negrita = Font(bold=True, size=20,  color='000000')
    # hoja.merge_cells('A1:D1')
    ws.append((titulo,))
    ws.append(("",))
    results.insert(0, nombre_filas)

    for row in results:
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    

    for celda in ws[1]:
        celda.font = negrita

    ws.merge_cells('A1:K1')

    wb.save(f"excel/{nombre_excel}.xlsx")

    return FileResponse(f"excel/{nombre_excel}.xlsx")




# def objetos_a_tuplas(objetos, atributos):
#     """
#     Convierte un array de objetos en una lista de tuplas con los atributos especificados.

#     :param objetos: Array de objetos a convertir.
#     :param atributos: Lista de nombres de atributos a extraer.
#     :return: Lista de tuplas con los valores de los atributos.
#     """
#     tuplas = []
#     for obj in objetos:
#         tupla = []
#         for atributo in atributos:
#             valor = getattr(obj, atributo, None)
#             if atributo in ['Verificado', 'Recibido']:  # Asumiendo que 'Verificado' y 'Recibido' son booleanos
#                 valor = cambiar_bool(valor)
#             tupla.append(valor)
#         tuplas.append(tuple(tupla))
#     return tuplas


def objetos_a_tuplas(objetos):
    """
    Convierte un array de objetos en una lista de tuplas con todos los atributos del objeto.

    :param objetos: Array de objetos a convertir.
    :return: Lista de tuplas con los valores de los atributos.
    """
    tuplas = []
    for obj in objetos:
        tupla = []
        atributos = vars(obj)
        for atributo, valor in atributos.items():
            if isinstance(valor, bool):  # Asumiendo que 'Verificado' y 'Recibido' son booleanos
                valor = cambiar_bool(valor)
            tupla.append(valor)
        tuplas.append(tuple(tupla))
    return tuplas

