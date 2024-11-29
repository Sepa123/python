import os
from fastapi import APIRouter, File, UploadFile, status,HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font , PatternFill, Border ,Side, Alignment
from datetime import datetime, timedelta
from openpyxl.worksheet.page import PageMargins , PrintPageSetup
from database.schema.operaciones.lista_vehiculos import vehiculos_disponibles_op_schema
import lib.excel_generico as excel
import time
import re
from typing import Dict, List
from database.schema.rutas.datos_rutas_tracking import datos_rutas_tracking_schema
from geopy.geocoders import Nominatim
import pandas as pd
## conexion

from database.client import reportesConnection , UserConnection
from database.hela_prod import HelaConnection

## modelos y schemas

from database.models.asignar_ruta import RutasAsignadas
from database.models.recepcion.recepcion_tiendas import bodyUpdateVerified
from database.models.ruta_manual import RutaManual
from database.schema.ruta_manual import convert_to_json, rutas_manuales_schema
from database.models.ruta_en_activo import RutaEnActivo
from database.schema.rutas.factura_electrolux import facturas_electrolux_schema
from database.schema.rutas_en_activo import rutas_en_activo_schema , ruta_en_activo_excel_schema
from database.schema.nombres_rutas_activas import nombres_rutas_activas_schema, comunas_ruta_schema, comuna_region_rutas_schema

from database.schema.rutas.cantidad_productos_rutas_activas import cant_productos_rutas_schema

from database.schema.rutas.alertas_conductor_ruta_activa import alertas_conductor_ruta_schema

from database.schema.datos_ruta_activa_editar import datos_rutas_activas_editar_schema
from database.schema.rutas.driver_ruta_asignada import driver_ruta_asignada

from database.schema.rutas.recuperar_tracking_beetrack import recuperar_tracking_beetrack_schema, recuperar_fecha_ingreso_sistema_schema

from database.schema.rutas.linea_producto import recuperar_linea_producto_schema

from database.schema.geolocalizacion.latlng import latlng_schema
from database.models.geolocalizacion.latlong import Latlong

from database.schema.rutas.archivo_descarga_beetrack import datos_descarga_beetracks_schema

from database.models.rutas.armar_rutas import ArmarRutaBloque

from database.models.rutas.lista_eliminar import ListaEliminar
from database.schema.rutas.buscar_producto_ruta import buscar_productos_ruta_schema
from database.schema.rutas.nombre_rutas_activa import nombre_rutas_activas_schema
from database.schema.rutas.reporte_ruta_mes import reportes_rutas_mes_schema, reporte_ruta_dia_schema
from database.schema.rutas.bitacora_log_inversa import bitacora_log_inversa_schema
from database.schema.rutas.no_entregados import no_entregados_schema

from database.schema.rutas.eficiencia_coductor import eficiencia_conductor_schema

from database.schema.rutas.seguimiento_ruta import seguimiento_ruta_schema
from database.schema.rutas.comuna_por_ruta import comunas_por_ruta_schema
from database.models.recepcion.recepcion_tiendas import bodyUpdateVerified
from database.schema.rutas.codigo_obligatorio_dia import codigos_obligatorios_dia_schema

router = APIRouter(tags=["rutas"], prefix="/api/rutas")

conn = reportesConnection()
connUser = UserConnection()
connHela = HelaConnection()

@router.post("/buscar",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_manual(body : bodyUpdateVerified ):

    results = conn.get_ruta_manual(body.n_guia)
    # print(results[0][10])

    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    

    check_producto  = results[0][10]
    # print(body)


    ### para pedidos pickeados de opl p
    if len(body.n_guia) > 20:
        cod_opl = conn.get_codigo_pedido_opl(body.n_guia)[0][0]
        body.n_guia = cod_opl

    # print(body)

    # print(results)

    check = conn.check_producto_existe(check_producto)
    check = re.sub(r'\(|\)', '',check[0])
    check = check.split(",")
    
    if(check[0] == "1"):
        print("codigo pedido repetido")
        check_fecha = conn.check_fecha_ruta_producto_existe(check_producto)

        if check_fecha is not None:
            fecha = check_fecha[0]
        else:
            fecha = ''
        
        raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                            detail=f'El Producto "{check_producto}" se encuentra en la ruta: {check[1]}, con fecha de ruta {fecha}' )
    
    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = rutas_manuales_schema(results)

    if json_data[0]['Calle'] is None:
        # print("La direccion es null")
        json_data[0]['Calle'] = json_data[0]['Direccion_textual']

    body.cliente = json_data[0]['Notas']

    data = body.dict()

    connHela.insert_data_bitacora_recepcion(data)

    return json_data

@router.post("/buscar/producto/ruta",status_code=status.HTTP_202_ACCEPTED)
async def get_datos_producto_en_ruta(body : bodyUpdateVerified ):
 
    codigo_ruta = conn.get_codigo_pedido_opl(body.n_guia)
    body.n_guia = codigo_ruta[0][0]

    results = conn.get_datos_producto_en_ruta(body.n_guia)
            
    if results is None or results == []:
        results = conn.get_datos_producto_en_ruta(body.cod_producto)
        if results is None or results == []:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = buscar_productos_ruta_schema(results)

    data = body.dict()

    connHela.insert_data_bitacora_recepcion(data)

    return json_data

@router.get("/buscar/rutas/activas",status_code=status.HTTP_202_ACCEPTED)
async def buscar_rutas_activas():
    results = conn.get_rutas_activas()
    return nombre_rutas_activas_schema(results)

@router.get("/buscar/{pedido_id}",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_manual(pedido_id : str):
    results = conn.get_ruta_manual(pedido_id)

    # print("codigo_pedido ",pedido_id)

    check = conn.check_producto_existe(pedido_id)
    check = re.sub(r'\(|\)', '',check[0])
    check = check.split(",")

    # print(check)
    
    if(check[0] == "1"):
        print("codigo pedido repetido")

        check_fecha = conn.check_fecha_ruta_producto_existe(pedido_id)

        if check_fecha is not None:
            fecha = check_fecha[0]
        else:
            pedido_id = conn.get_codigo_pedido_opl(pedido_id)[0][0]
            fecha = conn.check_fecha_ruta_producto_existe(pedido_id)[0]
        
        raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                            detail=f'El Producto "{pedido_id}" se encuentra en la ruta: {check[1]}, con fecha de ruta {fecha}' )
    
    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = rutas_manuales_schema(results)

    if json_data[0]['Calle'] is None:
        print("La direccion es null")
        json_data[0]['Calle'] = json_data[0]['Direccion_textual']
    # print(results)
    # print("/buscar/ruta")

    return json_data

@router.get("/buscar/sin_filtro/{pedido_id}",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_manual(pedido_id : str):
    
    results = conn.get_ruta_manual(pedido_id)

    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = rutas_manuales_schema(results)

    if json_data[0]['Calle'] is None:
        print("La direccion es null")
        json_data[0]['Calle'] = json_data[0]['Direccion_textual']
    # print(results)
    print("/buscar/ruta")

    return json_data

@router.get("/buscar/tracking/{pedido_id}",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_manual(pedido_id : str):
    
    results = conn.get_ruta_tracking_producto(pedido_id)

    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = datos_rutas_tracking_schema(results)

    if json_data[0]['Calle'] is None:
        print("La direccion es null")
        json_data[0]['Calle'] = json_data[0]['Direccion_textual']
    # print(results)
    print("/buscar/ruta")

    return json_data


@router.get("/buscar/factura/electrolux/{pedido_id}",status_code=status.HTTP_202_ACCEPTED)
async def get_factura_electrolux(pedido_id : str):
    # if re.search(r'BLE', pedido_id):
    #     factura = conn.get_numero_guia_by_factura(pedido_id)[0]
    #     result = conn.get_factura_electrolux(factura)
    # else:
    result = conn.get_factura_electrolux(pedido_id)

    return facturas_electrolux_schema(result)

def validar_fecha(fecha):
    fecha_actual = datetime.now().date()  # Obtiene la fecha actual
    
    # Convierte la fecha dada a un objeto de tipo datetime
    try:
        fecha_ingresada = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return False  # La fecha no tiene el formato correcto
    
    # Compara las fechas
    if fecha_ingresada >= fecha_actual:
        return True
    else:
        return False

@router.post("/agregar",status_code=status.HTTP_201_CREATED)
async def insert_ruta_manual(rutas : List[List[RutaManual]], fecha_pedido : str):
    try:
        # print(len(rutas))

        if validar_fecha(fecha_pedido) == False: raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
            detail=f"No se puede crear ruta con la fecha {fecha_pedido}")

        # print(fecha_pedido)
        id_ruta = conn.read_id_ruta()[0]
        nombre_ruta = conn.get_nombre_ruta_manual(rutas[0][0].Created_by)[0][0]

        check = conn.check_producto_existe(rutas[0][0].Codigo_pedido)
        check = re.sub(r'\(|\)', '',check[0])
        check = check.split(",")

        # print(check)

        if(check[0] == "1"):
            # print("codigo pedido repetido")
            # print(f'El Producto "{rutas[0][0].Codigo_pedido}" se encuentra en la ruta {check[1]}')

            raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                                detail=f'El Producto "{rutas[0][0].Codigo_pedido}" se encuentra en la ruta {check[1]}')
        for i, ruta in enumerate(rutas):
            for j,producto in enumerate(ruta):
                data = producto.dict()
                # print(data["Codigo_pedido"])
                data["Calle"] = conn.direccion_textual(data["Codigo_pedido"])[0][0]
                data["Id_ruta"] = id_ruta
                data["Agrupador"] = nombre_ruta
                data["Nombre_ruta"] = nombre_ruta
                # data["Pistoleado"]  
                data["Descripcion_producto"] = re.sub(r'@', ' ', producto.Descripcion_producto)
                data["Posicion"] = i + 1
                data["Fecha_ruta"] = fecha_pedido
                data["Pickeado"] = data["Pistoleado"] 
                if data["Fecha_ruta"] is None:
                    # Obtener la fecha actual
                    fecha_actual = datetime.now().date()
                    # Obtener la fecha del día siguiente
                    fecha_siguiente = fecha_actual + timedelta(days=1)
                    data["Fecha_ruta"] =fecha_siguiente
                conn.write_rutas_manual(data)
                # Ejecutar función en el último elemento
                if i == len(rutas) - 1 and j == len(ruta) - 1:
                    conn.recalcular_posicion_ruta(nombre_ruta)

                # if id_ruta == 1 and nombre_ruta == '':
                #     id_ruta = conn.read_id_ruta()[0]
                #     nombre_ruta = conn.get_nombre_ruta_manual(rutas[0][0].Created_by)[0][0]
                #     conn.update_id_ruta_rutas_manuales(id_ruta,nombre_ruta, data["Codigo_pedido"])


        
        return { "message": f"La Ruta {nombre_ruta} fue guardada exitosamente" }
    except Exception as e:
        if validar_fecha(fecha_pedido) == False: 
            raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, detail=f"No se puede crear ruta con la fecha {fecha_pedido}")
       
        if(check[0] == "1"):

            raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                                detail=f'El Producto "{rutas[0][0].Codigo_pedido}" se encuentra en la ruta {check[1]}')
        
        print(f"error al crear ruta: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
    # finally:
    #     conn.recalcular_posicion_ruta(nombre_ruta)
        

@router.put("/actualizar/estado/{cod_producto}",status_code=status.HTTP_202_ACCEPTED)
async def update_estado_producto(cod_producto:str, body : bodyUpdateVerified ):
     try:
          data = body.dict()
          conn.update_verified(cod_producto)
          connHela.insert_data_bitacora_recepcion(data)
          return { "message": "Producto actualizado correctamente" }
     except:
          print("error en /actualizar/estado/")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
     
@router.get("/listar/activo",status_code=status.HTTP_200_OK)
async def get_rutas_en_activo(nombre_ruta : str):
     print(nombre_ruta)
    #  results = conn.read_rutas_en_activo(nombre_ruta)
     results = conn.read_rutas_en_activo_para_armar_excel(nombre_ruta)

     if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="la ruta no existe")
     
    #  return rutas_en_activo_schema(results)
     return ruta_en_activo_excel_schema(results)


@router.post("/recuperar/bultos/sku",status_code=status.HTTP_200_OK)
async def recuperar_sku_productos_ruta(body : Dict):
     
     datos= []
     for code in body['codigos']:
        results = conn.prueba_recupera_bulto_sku(code,body['nombre_ruta'])
        # json =  {
        #     "guia" : results[0],
        #     "sku" : results[1],
        #     "descripcion" : results[2],
        #     "cant_producto" : results[3],
        #     "cant_x_producto" : results[4],
        #     "bultos" : results[5]
        # }
        # datos.append(json)

        datos.append(results)

     if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="la ruta no existe")

     return datos

## cantidad de productos (relacionados a read_ruta_en_activo)

@router.get("/listar/activo/cantidad/productos",status_code=status.HTTP_200_OK)
async def get_rutas_en_activo(nombre_ruta : str):
    #  print(nombre_ruta)
     results = conn.read_cant_productos_ruta_activa(nombre_ruta)

     if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="la ruta no existe")
     
     return cant_productos_rutas_schema(results)

# rutas activas

def separar_por_codigo_pedido(lista_objetos):
    # Crear un diccionario para almacenar los arrays según el Codigo_pedido
    arrays_por_pedido = {}

    # Iterar sobre cada objeto en la lista
    for objeto in lista_objetos:
        # Obtener el valor del Codigo_pedido
        codigo_pedido = objeto.Codigo_pedido

        # Verificar si ya hay un array para ese Codigo_pedido, si no, crear uno
        if codigo_pedido not in arrays_por_pedido:
            arrays_por_pedido[codigo_pedido] = []

        # Agregar el objeto al array correspondiente
        arrays_por_pedido[codigo_pedido].append(objeto)

    # Convertir el diccionario a una lista de arrays
    resultado = list(arrays_por_pedido.values())

    return resultado



@router.post("/agregar/ruta_activa_existente",status_code=status.HTTP_201_CREATED)
async def insert_ruta_existente_activa(fecha_ruta_nueva : str, rutas : List[List[RutaManual]]):
    try:
        fecha_ruta = fecha_ruta_nueva

        for i, ruta in enumerate(rutas):

            for producto in ruta:
                if producto.Id_ruta is not None:
                    id_ruta = producto.Id_ruta 
                    nombre_ruta = producto.Nombre_ruta 
                    break

            # valores_unicos = set(objeto.Codigo_pedido for objeto in ruta)
            # lista_val = list(valores_unicos)

            # if len(lista_val) > 1:
            #     nueva_pos = separar_por_codigo_pedido(ruta)
            #     index_eliminar = i
            #     rutas.pop(index_eliminar)
            #     for c,pos in enumerate(nueva_pos):
            #         rutas.insert(index_eliminar + c, pos )

            
        # id_ruta = rutas[0][0].Id_ruta
        # nombre_ruta = rutas[0][0].Nombre_ruta

        for i,ruta in enumerate(rutas):
            for j,producto in enumerate(ruta):
                producto.Peso=1
                producto.Volumen=1,
                producto.Dinero=1,
                producto.Duracion_min=8,
                producto.Ventana_horaria_1='09:00 - 21:00',
                producto.Ventana_horaria_2='',
                producto.Email_remitentes='',
                producto.Eliminar_pedido='',
                producto.Vehiculo='',
                producto.Habilidades=''

                data = producto.dict()

                # direccion_textual = conn.direccion_textual(data["Codigo_pedido"])
                data["Posicion"] = i + 1
                # print(i)
                check = conn.check_producto_codigo_repetido(nombre_ruta,data["Codigo_pedido"],data["Codigo_producto"], data["SKU"])
                # print("Codigo pedido",data["Codigo_pedido"])
                if check is not None:
                    data["Pickeado"] = data["Pistoleado"] 
                    if data["Pickeado"] == '1':
                        data["Pickeado"] = True
                    else: 
                        data["Pickeado"] = False
                    count = conn.update_posicion(data["Posicion"], data["Codigo_pedido"], data["Codigo_producto"], fecha_ruta, data["DE"], data["DP"], nombre_ruta, data["Pickeado"])
                    # print("Posicion:",data["Posicion"])
                else :
                    data["Calle"] = conn.direccion_textual(data["Codigo_pedido"])[0][0]
                    data["Id_ruta"] = id_ruta
                    data["Agrupador"] = nombre_ruta
                    data["Nombre_ruta"] = nombre_ruta
                    data["Pickeado"] = data["Pistoleado"] 
                    data["Fecha_ruta"] = fecha_ruta
                    # conn.update_verified(data["Codigo_producto"])
                    # print('posicion :',data["Posicion"])
                    conn.write_rutas_manual(data)

                if i == len(rutas) - 1 and j == len(ruta) - 1:
                    conn.recalcular_posicion_ruta(nombre_ruta)
        
        
        # conn.recalcular_posicion_ruta(nombre_ruta)
        return { "message": f"La Ruta {nombre_ruta} fue actualizada exitosamente" }
    except Exception as e:
        print("error an actualizar ruta: ",e)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
    # finally:
    #     conn.recalcular_posicion_ruta(nombre_ruta)

  
@router.get("/activo/nombre_ruta")
async def get_nombres_ruta(fecha : str):
    # read_nombres_rutas_comunas v1
    results = conn.read_nombres_rutas(fecha)
    return nombres_rutas_activas_schema(results)


@router.get("/activo/comunas")
async def get_nombres_ruta_comuna(fecha : str):

    # results = conn.read_comunas_ruta_by_fecha(fecha)
    # return comunas_ruta_schema(results)
    results = conn.read_comunas_regiones_ruta()
    return comuna_region_rutas_schema(results)


@router.get("/activo/nombre_ruta/filtro")
async def filter_nombre_ruta_by_comuna(fecha: str, comuna : str, region : str):
    
    if comuna == 'Todas' or comuna == 'Comunas' or comuna == '':
        results = conn.filter_nombres_rutas_by_region(fecha,comuna,region)
        # print("Filtro por region")
    else : 
        results = conn.filter_nombres_rutas_by_comuna(fecha,comuna,region)
        # print("Filtro por comuna")

    return nombres_rutas_activas_schema(results)

@router.put("/actualizar/estado/activo/{nombre_ruta}",status_code=status.HTTP_202_ACCEPTED)
async def update_estado_ruta(nombre_ruta:str, body : bodyUpdateVerified):
     try:
        #   print(nombre_ruta)
          data = body.dict()
          conn.update_estado_rutas(nombre_ruta)

          connHela.insert_data_bitacora_recepcion(data)
          return { "message": "Estado de Ruta Actualizado Correctamente" }
     except:
          print("error en /actualizar/estado/activo/nombre_ruta")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")

@router.get("/datos_ruta/{nombre_ruta}",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_by_nombre_ruta(nombre_ruta: str):
    #  print(nombre_ruta)
     results = conn.read_ruta_activa_by_nombre_ruta(nombre_ruta)
     return datos_rutas_activas_editar_schema(results)

@router.delete("/eliminar/producto/{cod_producto}/{nombre_ruta}") 
async def delete_producto_ruta_activa(cod_producto : str, nombre_ruta : str):
     try:
          results = conn.delete_producto_ruta_activa(cod_producto, nombre_ruta)
          if (results == 0): print("El producto no existe en ninguna ruta")
          return { "message" : "producto eliminado"}
     except:
          print("error en /eliminar/producto/cod_producto/nombre_ruta")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
     

## eliminar producto + bitacora
@router.put("/eliminar/producto/{cod_producto}") 
async def delete_producto_ruta_activa_bitacora(lista : ListaEliminar):
     try:
          results = conn.delete_producto_ruta_activa(lista.cod_producto, lista.nombre_ruta)
          if (results == 0): 
              print("El producto no existe en ninguna ruta")
          else:
              data = lista.dict()
              connHela.insert_data_bitacora_recepcion(data)
          return { "message" : "producto eliminado"}
     except:
          print("error en /eliminar/producto/cod_producto/nombre_ruta")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")    
     
## eliminar cargas
@router.put("/eliminar/productos")
async def delete_cargas_rsv(lista : ListaEliminar):

    if lista.lista == '':
        return {
        "message" : "no hay nada que eliminar",
        "mostrar" : True
    }
    codigos = lista.lista.split(',')

    # print(codigos)

    results = conn.delete_productos_ruta(lista.lista, lista.nombre_ruta)

    data = lista.dict()
    # connHela.insert_data_bitacora_recepcion(data)
    return {
        "message" : f"productos eliminados ,{results}",
        "mostrar" : True
    }

@router.delete("/eliminar/ruta/{nombre_ruta}") 
async def eliminar_ruta(nombre_ruta : str):
     try:
          results = conn.eliminar_ruta(nombre_ruta)
          if (results == 0): print("La ruta no existe en ninguna parte")
          return { "message" : "Ruta eliminada correctamente"}
     except:
          print("error en /eliminar/producto/")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
       
@router.post("/descargar/{var_random}")
async def download_excel(nombre_ruta : str,patente: str,driver:str , body : list, var_random : str,despachador : str):
    if despachador is None or despachador == 'null': 
        despachador = ""
    fecha_de_asignacion = conn.get_fecha_asignacion_ruta(nombre_ruta)[0]
    fecha_hoy = datetime.now()

    fecha_hoy =  fecha_hoy + timedelta(hours=1)

    fecha_impresion = fecha_hoy.strftime('%d-%m-%Y %H:%M:%S')

    fecha_asignacion = fecha_de_asignacion
    datos = [[]]
    
    datos.append([
        "N°", "Pedido", "Comuna","Nombre","Direccion", "Teléfono", "SKU", "Producto",
        "UND", "Bult","Obs"
    ])
  
    # result = conn.read_rutas_en_activo(nombre_ruta) 
    result = conn.read_rutas_en_activo_para_armar_excel(nombre_ruta)
    
    border = Border(left=Side(border_style='thin', color='000000'),   
                right=Side(border_style='thin', color='000000'),   
                top=Side(border_style='thin', color='000000'),     
                bottom=Side(border_style='thin', color='000000')) 
    
    rutas_activas = ruta_en_activo_excel_schema(result)
    # rutas_activas = body

    # Crear un libro de Excel y seleccionar la hoja activa
    libro_excel = Workbook()
    hoja = libro_excel.active
    hoja.title = 'Hoja1'    

    margins = PageMargins(top=0.3, bottom=0.6, left=0.4, right=0.5, header=0.3, footer=0.3)
    hoja.page_margins = margins

    # Estilo para el texto en negrita
    negrita = Font(bold=True, size=20,  color='000000')
    # hoja.merge_cells('A1:D1')
    hoja.append(("Ruta : "+nombre_ruta,))
    hoja.append(("Patente : "+patente,))
    hoja.append(("Fecha Asignación : "+ fecha_asignacion,))
    hoja.append(("Fecha Impresión : "+ fecha_impresion,))

    # "Patente : "+patente, "driver : "+driver

    for ruta in rutas_activas:
        arrayProductos = ruta["Producto"].split("@")
        arraySKU = ruta["SKU"].split("@")
        arrayUnidades = ruta["Unidades"].split("@")
        arrayBultos = ruta["Bultos"].split("@")
        if len(arraySKU) != len(arrayProductos):
            for i in range(len(arrayProductos)):
                arraySKU.append("")

        # print(ruta["arrayBultos"])
        if len(arrayProductos) == 1:
            fila = [
                ruta["Pos"], ruta["Codigo_pedido"], ruta["Comuna"] ,ruta["Nombre_cliente"],ruta["Direccion_cliente"], ruta["Telefono"], arraySKU[0], arrayProductos[0],
                arrayUnidades[0], arrayBultos[0] , ruta["DE"] + " " + ruta["DP"]
            ]
            datos.append(fila)
        elif len(arrayProductos) > 1:
            for i, producto in enumerate(arrayProductos):
                if i == 0:
                    fila = [
                        ruta["Pos"], ruta["Codigo_pedido"], ruta["Comuna"] ,ruta["Nombre_cliente"],ruta["Direccion_cliente"], ruta["Telefono"], arraySKU[0], producto,
                        arrayUnidades[0], arrayBultos[0], ruta["DE"] + " " + ruta["DP"]
                    ]
                    datos.append(fila)
                else:
                    fila_producto = [
                        "", "", "", "",
                        "", "", arraySKU[i], producto , arrayUnidades[i], arrayBultos[i]
                    ]
                    datos.append(fila_producto)
  
    # Escribir los datos en la hoja
    for i,fila in enumerate(datos):
        hoja.append(fila)
        nHoja = i
    
    # Aplicar estilo en negrita a la primera fila
    for celda in hoja[1]:
        celda.font = negrita

    for celda in hoja[2]:
        celda.font = negrita

    for celda in hoja[6]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="000000FF", end_color="000000FF", fill_type="solid")
        celda.border = border
    
    # aplicar largo a celdas

    for col_letter in ['A']:
      hoja.column_dimensions[col_letter].width = 3

    for col_letter in ['I','J']:
      hoja.column_dimensions[col_letter].width = 4

    for col_letter in ['B','G']:
      hoja.column_dimensions[col_letter].width = 12

    for col_letter in ['C','F']:
      hoja.column_dimensions[col_letter].width = 14
    
    for col_letter in ['K']:
      hoja.column_dimensions[col_letter].width = 15

    for col_letter in ['D']:
      hoja.column_dimensions[col_letter].width = 20

    for col_letter in ['E']:
      hoja.column_dimensions[col_letter].width = 25
    
    for col_letter in ['H']:
      hoja.column_dimensions[col_letter].width = 30

    # print(nHoja)
    for n in range(nHoja):
         for celda in hoja[n+7]:
             celda.font = Font(color="070707")
             celda.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
             celda.border = border
  
    hoja.append(("",)) 
    hoja.append(("","Driver : "+driver,'','','','Despachador: '+despachador,))  
    hoja.append(("",))  
    hoja.append(("","Firma : ______________________"))    

    hoja.merge_cells('A1:M1')
    hoja.merge_cells('A2:M2')

    for row in hoja.iter_rows(min_row=5, max_row=nHoja+5, min_col=0, max_col=13):
        for celda in row:
            celda.alignment = Alignment(wrap_text=True, horizontal = 'center' , vertical='center')

  # Fusionar celdas para las últimas cuatro filas
    # Guardar el archivo

    hoja.page_setup.orientation = 'landscape'

    # hoja.print_options.horizontalCentered = True  # Centrar horizontalmente
    # hoja.print_options.verticalCentered = True  # Centrar verticalmente

    nombre_archivo = "nombre_ruta.xlsx"
    libro_excel.save(nombre_archivo)

    headers = {
        "Cache-Control": "no-store, max-age=0",
        "Pragma": "no-cache",
    }

    return FileResponse(path="nombre_ruta.xlsx" ,headers=headers)

@router.post("/descargar")
async def download_excel_antigua(nombre_ruta : str,patente: str,driver:str , body : list):

    datos = [[]]
    
    datos.append([
        "N°", "Pedido", "Comuna","Nombre","Direccion", "Teléfono", "SKU", "Producto",
        "UND", "Bult","Obs"
    ])
  
    # result = conn.read_rutas_en_activo(nombre_ruta) 
    result = conn.read_rutas_en_activo_para_armar_excel(nombre_ruta)
    
    border = Border(left=Side(border_style='thin', color='000000'),   
                right=Side(border_style='thin', color='000000'),   
                top=Side(border_style='thin', color='000000'),     
                bottom=Side(border_style='thin', color='000000')) 
    
    rutas_activas = ruta_en_activo_excel_schema(result)
    

    # rutas_activas = body

    # Crear un libro de Excel y seleccionar la hoja activa
    libro_excel = Workbook()
    hoja = libro_excel.active
    hoja.title = 'Hoja1'    

    margins = PageMargins(top=0.3, bottom=0.6, left=0.4, right=0.5, header=0.3, footer=0.3)
    hoja.page_margins = margins

    # Estilo para el texto en negrita
    negrita = Font(bold=True, size=20,  color='000000')
    # hoja.merge_cells('A1:D1')
    hoja.append(("Ruta : "+nombre_ruta,))
    hoja.append(("Patente : "+patente,))

    # "Patente : "+patente, "driver : "+driver

    for ruta in rutas_activas:
        arrayProductos = ruta["Producto"].split("@")
        arraySKU = ruta["SKU"].split("@")
        arrayUnidades = ruta["Unidades"].split("@")
        arrayBultos = ruta["Bultos"].split("@")
        if len(arraySKU) != len(arrayProductos):
            for i in range(len(arrayProductos)):
                arraySKU.append("")

        # print(ruta["arrayBultos"])
        if len(arrayProductos) == 1:
            fila = [
                ruta["Pos"], ruta["Codigo_pedido"], ruta["Comuna"] ,ruta["Nombre_cliente"],ruta["Direccion_cliente"], ruta["Telefono"], arraySKU[0], arrayProductos[0],
                arrayUnidades[0], arrayBultos[0] , ruta["DE"] + " " + ruta["DP"]
            ]
            datos.append(fila)
        elif len(arrayProductos) > 1:
            for i, producto in enumerate(arrayProductos):
                if i == 0:
                    fila = [
                        ruta["Pos"], ruta["Codigo_pedido"], ruta["Comuna"] ,ruta["Nombre_cliente"],ruta["Direccion_cliente"], ruta["Telefono"], arraySKU[0], producto,
                        arrayUnidades[0], arrayBultos[0], ruta["DE"] + " " + ruta["DP"]
                    ]
                    datos.append(fila)
                else:
                    fila_producto = [
                        "", "", "", "",
                        "", "", arraySKU[i], producto , arrayUnidades[i], arrayBultos[i]
                    ]
                    datos.append(fila_producto)
  
    # Escribir los datos en la hoja
    for i,fila in enumerate(datos):
        hoja.append(fila)
        nHoja = i
    
    # Aplicar estilo en negrita a la primera fila
    for celda in hoja[1]:
        celda.font = negrita

    for celda in hoja[2]:
        celda.font = negrita

    for celda in hoja[4]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="000000FF", end_color="000000FF", fill_type="solid")
        celda.border = border
    
    # aplicar largo a celdas

    for col_letter in ['A']:
      hoja.column_dimensions[col_letter].width = 3

    for col_letter in ['I','J']:
      hoja.column_dimensions[col_letter].width = 4

    for col_letter in ['B','G']:
      hoja.column_dimensions[col_letter].width = 12

    for col_letter in ['C','F']:
      hoja.column_dimensions[col_letter].width = 14
    
    for col_letter in ['K']:
      hoja.column_dimensions[col_letter].width = 15

    for col_letter in ['D']:
      hoja.column_dimensions[col_letter].width = 20

    for col_letter in ['E']:
      hoja.column_dimensions[col_letter].width = 25
    
    for col_letter in ['H']:
      hoja.column_dimensions[col_letter].width = 30

    # print(nHoja)
    for n in range(nHoja):
         for celda in hoja[n+5]:
             celda.font = Font(color="070707")
             celda.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
             celda.border = border
  
    hoja.append(("",)) 
    hoja.append(("","Driver : "+driver,))  
    hoja.append(("",))  
    hoja.append(("","Firma : ______________________"))    

    hoja.merge_cells('A1:M1')
    hoja.merge_cells('A2:M2')

    for row in hoja.iter_rows(min_row=5, max_row=nHoja+5, min_col=0, max_col=13):
        for celda in row:
            celda.alignment = Alignment(wrap_text=True, horizontal = 'center' , vertical='center')

  # Fusionar celdas para las últimas cuatro filas
    # Guardar el archivo

    hoja.page_setup.orientation = 'landscape'

    # hoja.print_options.horizontalCentered = True  # Centrar horizontalmente
    # hoja.print_options.verticalCentered = True  # Centrar verticalmente

    nombre_archivo = "nombre_ruta.xlsx"
    libro_excel.save(nombre_archivo)

    return FileResponse("nombre_ruta.xlsx")

@router.post("/asignar")
async def asignar_ruta_activa(asignar : RutasAsignadas):
    try:
        asignar.id_ruta = conn.get_id_ruta_activa_by_nombre(asignar.nombre_ruta)[0]
        # print(asignar)
        data = asignar.dict()
        connHela.insert_ruta_asignada(data)

        return {"message": "ruta asignada correctamente"}
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error al ingresar la ruta ")
    
@router.get("/buscar_patente")
async def get_ruta_activa_by_nombre(nombre_ruta: str):
    try:
        results = connHela.read_id_ruta_activa_by_nombre(nombre_ruta)
        # print(results)
        if results is None:
            return { "OK": False}

        return driver_ruta_asignada(results)
    except:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error al ingresar la ruta ")

@router.put("/actualizar/ruta_asignada")
async def update_ruta_asignada(body :RutasAsignadas):
    try:
          connHela.update_ruta_asignada(body.patente,body.conductor,body.nombre_ruta , body.despachador)
          return { "message": "Ruta Actualizada Correctamente" }
    except:
          print("error en /actualizar/ruta_asignada")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")

@router.get("/beetrack/{id_ruta}/descargar/")
async def descargar_archivo_beetrack_antigua(id_ruta : str):
    results = conn.read_datos_descarga_beetrack(id_ruta)

    wb = Workbook()
    ws = wb.active
    print("/rutas/beetrack/descargar")
    # results.insert(0, ("",))
    results.insert(0, ("NÚMERO GUÍA *","VEHÍCULO *","NOMBRE ITEM *","CANTIDAD","CODIGO ITEM","IDENTIFICADOR CONTACTO *","NOMBRE CONTACTO", "TELÉFONO","EMAIL CONTACTO","DIRECCIÓN *",
    "LATITUD","LONGITUD","FECHA MIN ENTREGA","FECHA MAX ENTREGA","CT DESTINO","DIRECCION","DEPARTAMENTO","COMUNA","CIUDAD","PAIS","EMAIL","Fechaentrega","fechahr",
    "conductor","Cliente","Servicio","Origen","Región de despacho","CMN","Peso","Volumen", "Bultos","ENTREGA","FACTURA","OC","RUTA", "TIENDA"))

    for row in results:
        # print(row)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/prueba_beetrack.xlsx")

    return FileResponse("excel/prueba_beetrack.xlsx")

@router.get("/beetrack/{id_ruta}/descargar/{var_random}")
async def descargar_archivo_beetrack(id_ruta : str, var_random : str):
    # print("Esta es random ",var_random)
    results = conn.read_datos_descarga_beetrack(id_ruta)

    wb = Workbook()
    ws = wb.active
    print("/rutas/beetrack/descargar")
    # results.insert(0, ("",))
    results.insert(0, ("NÚMERO GUÍA *","VEHÍCULO *","NOMBRE ITEM *","CANTIDAD","CODIGO ITEM","IDENTIFICADOR CONTACTO *","NOMBRE CONTACTO", "TELÉFONO","EMAIL CONTACTO","DIRECCIÓN *",
    "LATITUD","LONGITUD","FECHA MIN ENTREGA","FECHA MAX ENTREGA","CT DESTINO","DIRECCION","DEPARTAMENTO","COMUNA","CIUDAD","PAIS","EMAIL","Fechaentrega","fechahr",
    "conductor","Cliente","Servicio","Origen","Región de despacho","CMN","Peso","Volumen", "Bultos","ENTREGA","FACTURA","OC","RUTA", "TIENDA"))

    for row in results:
        # print(row)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/prueba_beetrack.xlsx")

    headers = {
        "Cache-Control": "no-store, max-age=0",
        "Pragma": "no-cache",
    }

    return FileResponse(path="excel/prueba_beetrack.xlsx",headers=headers)

@router.get("/recuperar/tracking")
async def recuperar_tracking_beetrack(codigo : str):

    results = conn.recuperar_track_beetrack(codigo)
    # print(results)
    if results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Codigo no encontrado")
    
    return recuperar_tracking_beetrack_schema(results)

@router.get("/alerta/conductor",status_code=status.HTTP_202_ACCEPTED)
async def recuperar_alerta_ruta_activa(nombre_ruta :str):
    results = conn.get_alerta_carga_ruta_activa(nombre_ruta)
    return alertas_conductor_ruta_schema(results)

@router.get("/recuperar/linea/producto")
async def recuperar_linea_producto(codigo : str):

    results = conn.recupera_linea_producto(codigo)
    # print(results)
    if results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Codigo no encontrado")
    
    return recuperar_linea_producto_schema(results)
   
@router.get("/fecha_ingreso_sistema/{cod_pedido}")
async def recuperar_fecha_ingreso_sistema(cod_pedido : str):
    result = conn.recuperar_fecha_ingreso_cliente(cod_pedido)
    if(result == []) :
        #  print(f"codigo {cod_pedido} no encontrado")
         return {"Fecha_ingreso_sistema":"Sin Fecha de ingreso al sistema"}

    return recuperar_fecha_ingreso_sistema_schema(result)

@router.get("/tiempo")
async def test():
    time.sleep(90)
    return "hola"

@router.post("/geolocalizacion")
async def geolocalizar_direccion(body : Latlong):
    try:
        check = conn.check_direccion_existe(body.Direccion)
        # print(check[0])
        if check[0] >= 1:
            return f"La direccion {body.Direccion} ya se encuentra registrada"
        
        geolocalizacion = Nominatim(user_agent="backend/1.0")
        # ubicacion = geolocalizacion.reverse(f"{body.lat},{body.lng}",exactly_one=False)
        time.sleep(1)
        ubicacion = geolocalizacion.geocode(body.Direccion, exactly_one=True)

        if ubicacion is None :
            return f"No se encontro la ubicacion de {body.Direccion}"
        
        body.Lat = ubicacion.latitude
        body.Lng = ubicacion.longitude
        body.Display_name = ubicacion.address
        body.Type = ubicacion.raw['type']

        data = body.dict()
        conn.insert_latlng(data)

        return "direccion guardada en la tabla latlng "
    
    except:

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="hubo un error")
    

@router.post("/armar/bloque")
async def armar_rutas_predictivas(body : ArmarRutaBloque):
    data = body.dict()

    result = conn.armar_rutas_bloque(data)
    return result[0][1]

@router.get("/recalcular/ruta/{nombre_ruta}")
async def recalcular_posicion_rutas(nombre_ruta : str):
    # recalcular_posicion_ruta
    result = conn.recalcular_posicion_ruta(nombre_ruta)

    return {
        "message": "ruta recalculada"
    }



@router.get("/pedido/en_ruta/{pedido_id}")
async def pedido_en_ruta(pedido_id : str):
    check = conn.check_producto_existe(pedido_id)
    check = re.sub(r'\(|\)', '',check[0])
    check = check.split(",")

    return {
        "en_ruta" : check[0],
        "message": f"El pedido {pedido_id} no esta en ruta"
    }

@router.put("/actualizar/estado/activo/{nombre_ruta}/abrir",status_code=status.HTTP_202_ACCEPTED)
async def update_estado_ruta_a_true(nombre_ruta:str, body : bodyUpdateVerified):
     try:
        #   print(nombre_ruta)
          data = body.dict()
          conn.update_estado_rutas_a_true_abierta(nombre_ruta)
          connHela.insert_data_bitacora_recepcion(data)
          return { "message": "Estado de Ruta Actualizado Correctamente" }
     except:
          print("error en /actualizar/estado/activo/nombre_ruta/abrir ")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
     

@router.get("/reporte/mensual")
async def datos_rutas_mes(mes : str):
    result = conn.get_reportes_rutas_mes(mes)
    return reportes_rutas_mes_schema(result)



@router.get("/reporte/diario")
async def datos_rutas_mes(dia : str):
    result = conn.get_reportes_rutas_diario(dia)
    return reporte_ruta_dia_schema(result)


@router.get("/reporte/diario/excel/descargar")
async def datos_rutas_mes(dia : str):
    result = conn.get_reportes_rutas_diario(dia)
    nombre_filas = ( "Fecha Ruta", "Ruta", "Ruta Beetrack", "Posición", "Código Cliente", "Nombre",
                    "Dirección", "Comuna","Región","Telefono","Email","Código Pedido","Fecha pedido",
                    "Descripción","Cantidad","Creado por", "Daño embalaje", "Daño producto",
                    "Pickeado")
    nombre_excel = f"inventario-rutas-{dia}"

    return excel.generar_excel_generico(result,nombre_filas,nombre_excel)
    # return reporte_ruta_dia_schema(result)




##TRACKING LI

@router.get("/bitacora/log_inversa")
async def get_datos_logistica_inversa(cod_pedido : str):
     results = conn.get_bitacora_log_inversa_tracking(cod_pedido)
     pattern = r'\bportal-\b'

     bitacora_usuario = bitacora_log_inversa_schema(results)

     for usu in bitacora_usuario:
          if re.search(pattern,usu['Ids_usuario']) :
               id = usu['Ids_usuario'].replace("portal-","")
               nombre_usu = connUser.get_nombre_usuario(id)[0]
               usu['Nombre'] = nombre_usu
          else:
               id_hela = usu['Ids_usuario'].replace("hela-","")
               nombre_usu_hela = connHela.get_nombre_usuario_hela(id_hela)[0]
               usu['Nombre'] = nombre_usu_hela

          if usu['Observacion'] is None or usu['Observacion'] == '':
               usu['Observacion'] = "Sin observación"

     return bitacora_usuario


@router.post("/encontrar/producto/ruta",status_code=status.HTTP_202_ACCEPTED)
async def get_datos_producto_en_ruta(body : bodyUpdateVerified ):

    codigo_ruta = conn.get_codigo_pedido_opl(body.n_guia)

    body.n_guia = codigo_ruta[0][0]

    results = conn.get_datos_producto_en_ruta(body.n_guia)
            
    if results is None or results == []:
        results = conn.get_datos_producto_en_ruta(body.cod_producto)
        if results is None or results == []:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = buscar_productos_ruta_schema(results)

    data = body.dict()

    connHela.insert_data_bitacora_recepcion(data)

    return {
        "message" : "datos registrados en bitacora"
    }


@router.get("/no_entregados/total")
async def get_no_entregados_total(fecha : str,tienda : str, region: str):
    if tienda == 'undefined':
        tienda = None
        
    if region == 'undefined':
        region = None

    result = conn.read_no_entregados_total(fecha,tienda,region)
    return no_entregados_schema(result)



@router.get("/eficiencia/conductor")
async def get_eficiencia_conductor(fecha : str,tienda : str, region: str):
    if tienda == 'undefined':
        tienda = None
        
    if region == 'undefined':
        region = None

    result = conn.read_eficiencia_conductor(fecha,tienda,region)
    return eficiencia_conductor_schema(result)

@router.get("/media/eficiencia/conductor")
async def get_media_eficiencia_conductor(fecha : str,tienda : str, region: str):
    if tienda == 'undefined':
        tienda = None
        
    if region == 'undefined':
        region = None

    result = conn.read_media_eficiencia_conductor(fecha,tienda,region)
    return {
        "Suma": result[0],
        "T_ent": result[1],
        "N_ent": result[2],
        "Efectividad_entrega": result[3],
    }


@router.get("/regiones")
async def get_regiones():
    
    # result = conn.obtener_region()
    # obtener_region
    return [
        {
            "region" : "V - Valparaiso"
        },
        {
            "region" : "XIII - Metropolitana"
        }
    ]




@router.post("/agregar/porDespachoRuta",status_code=status.HTTP_201_CREATED)
async def insert_ruta_manual_por_despacho_ruta(body : bodyUpdateVerified):
    results = conn.get_ruta_manual(body.n_guia)
    # print(results[0][10])

    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    

    check_producto  = results[0][10]
    # print(body)


    ### para pedidos pickeados de opl p
    if len(body.n_guia) > 20:
        cod_opl = conn.get_codigo_pedido_opl(body.n_guia)[0][0]
        body.n_guia = cod_opl

    # print(body)

    # print(results)

    check = conn.check_producto_existe(check_producto)
    check = re.sub(r'\(|\)', '',check[0])
    check = check.split(",")
    
    if(check[0] == "1"):
        print("codigo pedido repetido")
        check_fecha = conn.check_fecha_ruta_producto_existe(check_producto)

        if check_fecha is not None:
            fecha = check_fecha[0]
        else:
            fecha = ''
        
        raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                            detail=f'El Producto "{check_producto}" se encuentra en la ruta: {check[1]}, con fecha de ruta {fecha}' )
    
    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = rutas_manuales_schema(results)

    if json_data[0]['Calle'] is None:
        print("La direccion es null")
        json_data[0]['Calle'] = json_data[0]['Direccion_textual']

    body.cliente = json_data[0]['Notas']

    data = body.dict()

    # print(json_data)
    # connHela.insert_data_bitacora_recepcion(data)

    try:
        for j,producto in enumerate(json_data):
            data = producto
            # print(data["Codigo_pedido"])
            data["Calle"] = conn.direccion_textual(data["Codigo_pedido"])[0][0]
            data["Id_ruta"] = body.id_ruta
            data["Agrupador"] = body.ruta
            data["Nombre_ruta"] = body.ruta
            # data["Pistoleado"]  
            data["Descripcion_producto"] = re.sub(r'@', ' ', producto['Descripcion_producto'])
            data["Posicion"] = j + 1
            data["Fecha_ruta"] = body.fecha_ruta
            data["Pickeado"] = data["Pistoleado"] 
            data["Created_by"] = body.id_usuario
            data["DE"] = False
            data["DP"] = False
            conn.write_rutas_manual(data)
            # Ejecutar función en el último elemento
            if j == len(json_data) - 1:
                conn.recalcular_posicion_ruta(body.ruta)


        
        return { "message": f"La Ruta {body.ruta} fue guardada exitosamente" }
    except Exception as e:
        # if validar_fecha(fecha_pedido) == False: 
        #     raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, detail=f"No se puede crear ruta con la fecha {fecha_pedido}")
       
        if(check[0] == "1"):

            raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                                detail=f'El Producto "{body.n_guia}" se encuentra en la ruta {check[1]}')
        
        print(f"error al crear ruta: {e}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
    # finally:
    #     conn.recalcular_posicion_ruta(nombre_ruta)

    # return body




@router.get("/buscar/4_digitos/{pedido_id}",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_manual(pedido_id : str):
    results = conn.buscar_productos_por_4_digitos(pedido_id)

    # print("codigo_pedido ",pedido_id)

    check = conn.check_producto_existe(pedido_id)
    check = re.sub(r'\(|\)', '',check[0])
    check = check.split(",")

    # print(check)
    
    if(check[0] == "1"):
        print("codigo pedido repetido")

        check_fecha = conn.check_fecha_ruta_producto_existe(pedido_id)

        if check_fecha is not None:
            fecha = check_fecha[0]
        else:
            pedido_id = conn.get_codigo_pedido_opl(pedido_id)[0][0]
            fecha = conn.check_fecha_ruta_producto_existe(pedido_id)[0]
        
        raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED, 
                            detail=f'El Producto "{pedido_id}" se encuentra en la ruta: {check[1]}, con fecha de ruta {fecha}' )
    
    if results is None or results == []:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    
    json_data = buscar_productos_ruta_schema(results)

 
    print("/buscar/ruta")

    return json_data



## seguimento ruta
@router.get("/seguimento")
async def get_seguimento_ruta():
    
    result = conn.seguimiento_transporte()
    # obtener_region
    return seguimiento_ruta_schema(result)



## seguimento ruta
@router.get("/comuna_por_ruta/descargar")
async def get_comuna_por_ruta(fecha : str):
    
    result = conn.comuna_por_ruta(fecha)
    # obtener_region
    nombre_filas = ( "Nombre Ruta", "Comunas", "Total Puntos", "Compromisos")
    nombre_excel = f"lista_comunas_por_ruta"

    return excel.generar_excel_generico(result,nombre_filas,nombre_excel)

## seguimento ruta
@router.get("/comuna_por_ruta")
async def get_comuna_por_ruta(fecha : str):
    result = conn.comuna_por_ruta(fecha)
    # obtener_region
    return comunas_por_ruta_schema(result)

## seguimento ruta
@router.get("/codigos/obligatorios/dia")
async def get_codigo_obligatorios_dia(fecha : str):
    result = conn.codigos_obligatorios_dia(fecha)
    # obtener_region
    return codigos_obligatorios_dia_schema(result)


## seguimento ruta
@router.get("/codigos/obligatorios/dia/descargar")
async def get_codigo_obligatorios_dia(fecha : str):
    result = conn.codigos_obligatorios_dia_excel(fecha)
    nombre_filas = ( "Clientes", "Código Pedido", "Fecha Pedido", "Fecha Reprogramada", "Comuna", "Región", "Descripción", "Sin Morador", "Verificado",
                    "Recepcionado", "Ruta Hela")
    nombre_excel = f"lista_codigo_obligatorios"

    return excel.generar_excel_generico(result,nombre_filas,nombre_excel)


## Vehiculos disponibles en ruta
@router.get("/vehiculos/disponibles/en_ruta")
async def get_veh_disp_operaciones(fecha : str):

    result = conn.obtener_veh_disp_operaciones(fecha)
    return vehiculos_disponibles_op_schema(result)


@router.get("/patentes/disponibles")
async def get_lista_patentes_disponibles():
    datos = conn.obtener_patentes_disponibles_crv_crm()
    return datos[0]



@router.post("/quadminds/subir-archivo", status_code=status.HTTP_202_ACCEPTED)
async def subir_archivo(id_usuario : str, file: UploadFile = File(...)):

    # select quadminds.convierte_en_ruta_manual(1,'202308021040');

    directorio  = os.path.abspath("excel")

    ruta = os.path.join(directorio,file.filename)

    with open(ruta, "wb") as f:
        contents = await file.read()
        print("pase por aqui")
        f.write(contents)

    df = pd.read_excel(ruta)

    df = df.applymap(lambda x: None if pd.isna(x) else x)

    lista = df.to_dict(orient='records')

    for i, data in enumerate(lista):
        # cantidad_encontrada = conn.get_pedido_planificados_quadmind_by_cod_pedido()
        # if cantidad_encontrada[0] >= 1:
        #     print("Producto ya esta registrado") 
        # else:
        if not data['razon_social'] or not data['domicilio'] or not data['fecha_reparto'] or not data['maquina'] or not data['fecha_pedido'] or not data['cod_pedido'] or not data['cod_producto'] or not data['producto'] or not data['cantidad']:
            print(f"Registro en la posición {i+1} ignorado por campos vacíos")
            continue  

        direccion = data['domicilio']
        posicion = i + 1
        conn.write_pedidos_planificados(data ,posicion, direccion)
        # print(posicion)


    fecha_hora_actual = datetime.now()

    fecha_dia = fecha_hora_actual.strftime("%Y%m%d")

    fecha_hora_formateada = fecha_hora_actual.strftime("%Y%m%d%H%M")
   
    error = conn.asignar_ruta_quadmind_manual(id_usuario, fecha_dia)

    diferencia = conn.calcular_diferencia_tiempo(fecha_dia)

    # error 1 : codigos inexistentes

    if error[0][0] == 1:
        return {
                "filename": file.filename, 
                "message": "Error al subir el archivo", 
                "codigos": f"{error[0][1]}",
                "tiempo": diferencia[0][0],
                "termino" : True ,
                "error" : 1,
                }
    else:   
        return {"filename": file.filename, 
                "message": error[0][1], 
                "codigos": "",
                "tiempo": diferencia[0][0],
                "termino" : True,
                "error" : 0,
                }
