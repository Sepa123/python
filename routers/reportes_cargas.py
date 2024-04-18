from fastapi import APIRouter,status,HTTPException
from database.client import reportesConnection
from fastapi.responses import FileResponse
from openpyxl import Workbook
import time 
import re ,json
from urllib.parse import unquote
# from os import remove
from database.models.producto_picking import producto_picking
from database.schema.producto_picking import producto_picking_schema, productos_picking_schema

from database.models.reporte_historico import ReporteHistorico
from database.schema.reporte_historico import reportes_historico_schema

from database.models.carga_easy import CargaEasy
from database.schema.cargas_easy import cargas_easy_schema

from database.schema.reporte_hora import reportes_hora_schema, reportes_ultima_hora_schema
from database.schema.reportes_easy_region import reportes_easy_region_schema

from database.models.reporte_productos_entregados import ReporteProducto
from database.schema.reporte_productos_entregados import reportes_producto_schema

from database.models.pedidos_compromiso_sin_despacho import PedidosCompromisoSinDespacho
from database.schema.pedidos_compromiso_sin_despacho import pedidos_compromiso_sin_despacho_schema

from database.schema.rutas_beetrack_hoy import rutas_beetrack_hoy_schema

from database.models.pedidos import Pedidos
from database.schema.pedidos import pedidos_schema

from database.models.pedidos_sin_tienda import PedidosSinTienda
from database.schema.pedidos_sin_tienda import pedidos_sin_tienda_schema

from database.models.pedidos_tienda_easy_opl import PedidosTiendaEasyOPL
from database.schema.pedidos_tiendas_easy_opl import pedidos_tiendas_easy_opl_schema

from database.models.pedidos_pendientes import PedidosPendientes
from database.schema.pedidos_pendientes import pedidos_pendientes_schema

from database.models.carga_easy_comparacion import CargaEasyComparacion
from database.schema.carga_easy_comparacion import cargas_easy_comparacion_schema


from database.models.operaciones.nro_cargas_hora import NroCargasHora
from database.schema.operaciones.nro_cargas_hora import nro_cargas_hora_schema

from database.schema.cargas.beetrack_rango import beetrack_rango_schema

from database.schema.pendientes_bodega import pendientes_bodega_schema

from database.models.ns_valor_ruta import asignarValor

from database.schema.estado.ns_verificado import ns_verificados_schema
from database.schema.nivel_servicio.ns_fecha_real import ns_por_fecha_schema , ns_por_fecha_inicial_schema
from database.schema.nivel_servicio.ns_drivers import ns_drivers_schema
from database.schema.nivel_servicio.ns_easy import ns_easy_schema,ns_pendientes_easy_region_schema , panel_principal_ns_easy,panel_regiones_ns_easy_schema, panel_noentregas_easy_schema

from database.models.nivel_servicio.ns_easy import  NSEasy
import lib.excel_generico as excel
import lib.guardar_datos_json as datos_json

import datetime
from typing import List
from fastapi.params import Query

router = APIRouter(tags=["reportes"],prefix="/api/reportes")

conn = reportesConnection()

@router.get("/cargas_easy",status_code=status.HTTP_202_ACCEPTED)
async def get_cuenta():
    data_db = conn.read_cargas_easy()

    if not data_db:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la consulta")
    
    return cargas_easy_schema(data_db)

### Quadminds
@router.get("/clientes",status_code=status.HTTP_202_ACCEPTED)
async def get_data_cliente():
    results = conn.read_clientes()
    # print(results)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("","","","",""))
    results.insert(1, ("Codigo de Cliente","Nombre","Calle y Número","Ciudad","Provincia/Estado","Latitud",
                       "Longitud","Teléfono con código de país","Email","Código de Pedido","Fecha de Pedido","Operación E/R",
                       "Código de Producto","Descripción del Producto","Cantidad de Producto","Peso","Volumen","Dinero",
                       "Duración min","Ventana horaria 1","Ventana horaria 2","Notas","Agrupador","Email de Remitentes","Eliminar Pedido Si - No - Vacío",
                       "Vehículo","Habilidades"))
    for row in results:
        texto = row[2]
        texto_limpio = re.sub(r'[\x01]', '', str(texto))
        new_row = row[:2] + (texto_limpio,) + row[3:]
        ws.append(new_row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save("Carga_Quadminds_yyyymmdd-hh24miss.xlsx")

    return FileResponse("Carga_Quadminds_yyyymmdd-hh24miss.xlsx")

@router.get("/clientes/json",status_code=status.HTTP_202_ACCEPTED)
async def get_data_cliente():
    results = conn.read_clientes()
    # print(len(productos_picking_schema(results)))
    return productos_picking_schema(results)

@router.get("/quadminds/fecha_compromiso",status_code=status.HTTP_202_ACCEPTED )
async def get_quadminds_fecha_compromiso():
    results = conn.read_reporte_quadmind_fecha_compromiso()
    # print(results)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("","","","",""))
    results.insert(1, ("Codigo de Cliente","Nombre","Calle y Número","Ciudad","Provincia/Estado","Latitud",
                       "Longitud","Teléfono con código de país","Email","Código de Pedido","Fecha de Pedido","Operación E/R",
                       "Código de Producto","Descripción del Producto","Cantidad de Producto","Peso","Volumen","Dinero",
                       "Duración min","Ventana horaria 1","Ventana horaria 2","Notas","Agrupador","Email de Remitentes","Eliminar Pedido Si - No - Vacío",
                       "Vehículo","Habilidades"))
    for row in results:
        texto = row[2]
        texto_limpio = re.sub(r'[\x01]', '', str(texto))
        new_row = row[:2] + (texto_limpio,) + row[3:]
        ws.append(new_row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save("Carga_Quadminds_Fecha_Compromiso_yyyymmdd-hh24miss.xlsx")

    return FileResponse("Carga_Quadminds_Fecha_Compromiso_yyyymmdd-hh24miss.xlsx")

@router.get("/quadminds/fecha_compromiso/json",status_code=status.HTTP_202_ACCEPTED)
async def get_data_cliente():
    results = conn.read_reporte_quadmind_fecha_compromiso()
    # print(len(productos_picking_schema(results)))
    return productos_picking_schema(results)

@router.get("/quadminds/tamano",status_code=status.HTTP_202_ACCEPTED )
async def get_quadminds_fecha_compromiso():
    results = conn.read_resumen_quadmind_tamano()
    # print(results)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("","","","",""))
    results.insert(1, ("Codigo de Cliente","Nombre","Calle y Número","Ciudad","Provincia/Estado","Latitud",
                       "Longitud","Teléfono con código de país","Email","Código de Pedido","Fecha de Pedido","Operación E/R",
                       "Código de Producto","Descripción del Producto","Cantidad de Producto","Peso","Volumen","Dinero",
                       "Duración min","Ventana horaria 1","Ventana horaria 2","Notas","Agrupador","Email de Remitentes","Eliminar Pedido Si - No - Vacío",
                       "Vehículo","Habilidades"))
    for row in results:
        texto = row[2]
        texto_limpio = re.sub(r'[\x01]', '', str(texto))
        new_row = row[:2] + (texto_limpio,) + row[3:]
        ws.append(new_row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save("excel/Carga_Quadminds_tamano_yyyymmdd-hh24miss.xlsx")

    return FileResponse("excel/Carga_Quadminds_tamano_yyyymmdd-hh24miss.xlsx")


@router.get("/NS_beetrack_Mensual",status_code=status.HTTP_202_ACCEPTED)
async def get_beetrack_mensual():
    results = conn.read_NS_beetrack_mensual()

    wb = Workbook()
    ws = wb.active
    results.insert(0, ("",))
    results.insert(1,('FECHA', 'ID. RUTA', 'DRIVER', 'PATENTE', 'REGION', 'Km. Ruta', 'T-PED', 'Easy', 'Electrolux', 'Sportex', 'Imperial', 'PBB', 'Virutex', 'R1', 'R2', 'R3', 'VR', 'C11', '(%) 11', 'C13', '(%) 13', 'C15', '(%)15', 'C17', '(%)17', 'C18', '(%)18', 'C20', '(%)20', 'Final_D', 'OBSERV-RUTA', 'H_INIC', 'H_TERM', 'TT-RUTA', 'Prom. ENT', 'T-ENT', 'N-ENT', 'EE', 'SM', 'CA', 'DA', 'RxD', 'DNE', 'DNCC', 'D.ERR', 'INC.T', 'DFORM', 'PINCOM', 'SPELI', 'PNCORR', 'PFALT', 'PPARC', 'P.DUPL', 'R', 'Pedidos'))

    for row in results:

        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/NS_Beetrack_Mensual.xlsx")

    return FileResponse("excel/NS_Beetrack_Mensual.xlsx")

@router.get("/NS_beetrack/rango",status_code=status.HTTP_202_ACCEPTED)
async def get_beetrack_rango(fecha_inicio: str, fecha_fin : str):
    results = conn.get_NS_beetrack_por_rango_fecha(fecha_inicio,fecha_fin)
    return beetrack_rango_schema(results)


#asignar valor a la ruta existente
@router.put("/NS_beetrack/rango",status_code=status.HTTP_202_ACCEPTED)
async def update_beetrack_valor_ruta(body: List[asignarValor]):
    if body == []:
         raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Error, no se enviaron datos")
    output = conn.update_valor_rutas(body)
    return { "message":f"Valor agregado correctamente "}


@router.get("/NS_beetrack/rango/descargar",status_code=status.HTTP_202_ACCEPTED)
async def get_beetrack_rango(fecha_inicio: str, fecha_fin : str):
    results = conn.get_NS_beetrack_por_rango_fecha(fecha_inicio,fecha_fin)
    # beetrack_dict = beetrack_rango_schema(results) 
    wb = Workbook()
    ws = wb.active
    
    results.insert(0, ("",))
    results.insert(1,("FECHA", "ID. RUTA","DRIVER","PATENTE","REGION","Km. Ruta","T-PED","Easy","Electrolux","Sportex","Imperial","PBB","Virutex","R1","R2","R3","VR",
    "C11","(%) 11","C13","(%) 13","C15","(%)15","C17","(%)17","C18","(%)18","C20","(%)20","Final_D","OBSERV-RUTA","H_INIC","H_TERM","TT-RUTA","Prom. ENT","T-ENT",
    "N-ENT","EE","SM","CA","DA","RxD","DNE","DNCC","D.ERR","INC.T","DFORM","PINCOM","SPELI","PNCORR","PFALT","PPARC","P.DUPL","R","Pedidos","Valor_ruta"
))
    for row in results:
        # print(row)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/NS_beetrack_rango.xlsx")

    return FileResponse("excel/NS_beetrack_rango.xlsx")

## Reportes Historicos
reporte_historico = None
ultima_ejecucion = None

def get_historico_mensual_unico():
    global reporte_historico
    results = conn.read_reporte_historico_mensual()
    reporte_historico = reportes_historico_schema(results)

def ejecutar_solo_una_vez_al_dia():
    global ultima_ejecucion 

    # Obtener la fecha y hora actual
    ahora = datetime.datetime.now()

    # Definir la hora a la que quieres que se ejecute la función (por ejemplo, a las 3 AM)
    hora_programada = datetime.time(3, 0, 0)

    # Combinar la fecha actual con la hora programada
    momento_programado = datetime.datetime.combine(ahora.date(), hora_programada)

    # Verificar si la función ya se ejecutó hoy
    if ultima_ejecucion is None or ahora - ultima_ejecucion > datetime.timedelta(hours=9):
        # Ejecutar la función
        get_historico_mensual_unico()

        # Actualizar el momento de la última ejecución a la hora actual
        ultima_ejecucion = ahora

        print("Ultima Ejecucion Reporte Historico", ultima_ejecucion)


@router.get("/historico/mensual",status_code=status.HTTP_202_ACCEPTED)
async def get_historico_mensual():
    ejecutar_solo_una_vez_al_dia()
    # results = conn.read_reporte_historico_mensual()
    # return reportes_historico_schema(results)
    return reporte_historico

@router.get("/historico/hoy",status_code=status.HTTP_202_ACCEPTED)
async def get_historico_hoy():
    results = conn.read_reporte_historico_hoy()
    return reportes_historico_schema(results)

@router.get("/historico/anual",status_code=status.HTTP_202_ACCEPTED)
async def get_historico_mensual():
    results = conn.read_reporte_historico_anual()

    wb = Workbook()
    ws = wb.active
    
    results.insert(0, ("",))
    results.insert(1,('Día', 'Fecha', 'Electrolux', 'Sportex', 'Easy', 'Tiendas'))
    for row in results:
        # print(row)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/Reporte_historico_mensual.xlsx")

    return FileResponse("excel/Reporte_historico_mensual.xlsx")

## reporte productos entregados

@router.get("/productos/mensual",status_code=status.HTTP_202_ACCEPTED)
async def get_productos_mensual():
    results = conn.read_reporte_producto_entregado_mensual()
    return reportes_producto_schema(results)

@router.get("/productos/hoy",status_code=status.HTTP_202_ACCEPTED)
async def get_productos_hoy():
    results = conn.read_reporte_producto_entregado_hoy()
    return reportes_producto_schema(results)

@router.get("/productos/rango",status_code=status.HTTP_202_ACCEPTED)
async def get_productos_por_rango_fecha(inicio, termino):
    if inicio == 'undefined' or termino == 'undefined':
    # Una o ambas variables son None
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="las varaibles no estan definidas")
  
    results = conn.read_reporte_producto_entregado_por_rango_fecha(inicio,termino)
    
    return reportes_producto_schema(results)

## TODO: EL ANUAL ES PARA DESCARGA
@router.get("/productos/anual",status_code=status.HTTP_202_ACCEPTED)
async def get_productos_anual():
    results = conn.read_reporte_producto_entregado_anual()
    return reportes_producto_schema(results)

# Cantidad de Entregas por hora
@router.get("/hora",status_code=status.HTTP_202_ACCEPTED)
async def get_reportes_hora():
    results = conn.read_reportes_hora()
    return reportes_hora_schema(results)

@router.get("/ultima_hora",status_code=status.HTTP_202_ACCEPTED)
async def get_reportes_ultima_hora():
    results = conn.read_reporte_ultima_hora()
    return reportes_ultima_hora_schema(results)


@router.get("/productos/easy_region",status_code=status.HTTP_202_ACCEPTED)
async def get_productos_easy_region():
    results = conn.read_productos_easy_region()
    return reportes_easy_region_schema(results)

@router.get("/pedidos/sin_despacho",status_code=status.HTTP_202_ACCEPTED)
async def get_pedidos_sin_despacho():
    results = conn.read_pedido_compromiso_sin_despacho()
    return pedidos_compromiso_sin_despacho_schema(results)

@router.get("/pedidos/sin_despacho/descargar",status_code=status.HTTP_202_ACCEPTED)
async def get_pedidos_sin_despacho_descarga():

    results = conn.read_pedido_compromiso_sin_despacho()

    wb = Workbook()
    ws = wb.active
    results.insert(0, ("",))
    results.insert(1,('Origen', 'Cod. Entrega', "Fecha Ingreso", "Fecha Compromiso", 
                      "Region", "Comuna","Descripcion","Bultos"))

    for row in results:
        # print(row)
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter# get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    results.insert(0, ("",))
    wb.save("excel/pedidos_con_fecha_de_compromiso_sin_despacho.xlsx")

    return FileResponse("excel/pedidos_con_fecha_de_compromiso_sin_despacho.xlsx")

@router.get("/pedidos",status_code=status.HTTP_202_ACCEPTED)
async def get_pedidos():
    results = conn.read_pedidos()
    return pedidos_schema(results)

@router.get("/ruta/beetrack/hoy",status_code=status.HTTP_202_ACCEPTED)
async def get_ruta_beetrack_hoy():
    results = conn.read_ruta_beetrack_hoy()
    return rutas_beetrack_hoy_schema(results)

@router.get("/pedidos/sin_tiendas",status_code=status.HTTP_202_ACCEPTED)
async def get_pedidos_sin_tienda():
    results = conn.read_pedidos_sin_tienda()
    return pedidos_sin_tienda_schema(results)

@router.get("/pedidos/easy_opl",status_code=status.HTTP_202_ACCEPTED)
async def get_pedidos_tiendas_easy_opl():
    results = conn.read_pedidos_tiendas_easy_opl()
    return pedidos_tiendas_easy_opl_schema(results)

@router.get("/pedidos/pendientes/total",status_code=status.HTTP_202_ACCEPTED)
async def pedidos_pendientes_total():
    results = conn.read_pedidos_pendientes_total()
    return pedidos_pendientes_schema(results)

@router.get("/pedidos/pendientes/entregados",status_code=status.HTTP_202_ACCEPTED)
async def pedidos_pendientes_total():
    results = conn.read_pedidos_pendientes_entregados()
    return pedidos_pendientes_schema(results)

@router.get("/pedidos/pendientes/no_entregados",status_code=status.HTTP_202_ACCEPTED)
async def pedidos_pendientes_total():
    results = conn.read_pedidos_pendientes_no_entregados()
    return pedidos_pendientes_schema(results)

@router.get("/pedidos/pendientes/en_ruta",status_code=status.HTTP_202_ACCEPTED)
async def pedidos_pendientes_total():
    results = conn.read_pedidos_pendientes_en_ruta()
    return pedidos_pendientes_schema(results)

## Comparacion cargas easy API VS WMS

@router.get("/cargas_easy/api",status_code=status.HTTP_202_ACCEPTED)
async def cargas_easy_api():
    results = conn.read_carga_easy_api()
    return cargas_easy_comparacion_schema(results)

@router.get("/cargas_easy/wms",status_code=status.HTTP_202_ACCEPTED)
async def cargas_easy_api():
    results = conn.read_carga_easy_wms()
    return cargas_easy_comparacion_schema(results)

#producto picking

@router.get("/buscar/producto")
async def productos_picking():
    results = conn.get_producto_picking()
    return producto_picking_schema(results)

@router.get("/buscar/producto/{producto_id}",status_code=status.HTTP_202_ACCEPTED)
async def producto_picking_id(producto_id : str):
    results = conn.get_producto_picking_id(producto_id)
    # print(results[""])
    if results is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="El codigo del producto no existe")
    return producto_picking_schema(results)

# cargas por hora

@router.get("/cargas_por_hora",status_code=status.HTTP_202_ACCEPTED)
async def cargas_por_hora():
    results = conn.get_carga_hora()

    return nro_cargas_hora_schema(results)


@router.get("/ns/verificados",status_code=status.HTTP_202_ACCEPTED)
async def Nivel_servicio_verificados(fecha_inicio : str,fecha_fin: str):
    results = conn.ns_picking(fecha_inicio,fecha_fin)
    return ns_verificados_schema(results)

#Pendientes en bodega
@router.get("/pendientes/bodega",status_code=status.HTTP_202_ACCEPTED)
async def cargas_por_hora():
    results = conn.read_lista_pendientes_bodega_hasta_hoy()

    return pendientes_bodega_schema(results)


### promedio ns fecha
@router.get("/ns/fecha_c_real",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_fecha_real(fecha : str):
    results = conn.revisar_nivel_servicio_fec_real(fecha)

    promedio = conn.revisar_nivel_servicio_fec_real_promedio(fecha)[0]

    return {
        "datos" : ns_por_fecha_inicial_schema(results),
        "promedio" : promedio
    }



### promedio ns fecha easy
@router.get("/ns/fecha_c_real/easy",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_fecha_real_easy(fecha : str, tienda: str):

    if tienda == 'easy':
        results = conn.revisar_nivel_servicio_fec_real_easy(fecha)

    if tienda == 'opl':
        results = conn.revisar_nivel_servicio_fec_real_easy_opl(fecha)

    if tienda == 'elux':
        results = conn.revisar_nivel_servicio_fec_real_elux(fecha)

    return {
        "datos" : ns_por_fecha_schema(results),
    }



### Nivel Servicio drivers
@router.get("/ns/drivers",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_drivers(fecha_inicio : str, fecha_fin: str):

    results = conn.nivel_servicio_drivers(fecha_inicio,fecha_fin)

    datos = ns_drivers_schema(results)

    patentes = [dato['Patente'] for dato in datos]
    porcentaje = [dato['P_ee'] for dato in datos]
    pedidos = [dato['Total_pedidos'] for dato in datos]
    Entregados = [dato['Total_entregados'] for dato in datos]


    return {
        'Datos' : datos,
        'Patentes' : patentes,
        'Porcentaje' : porcentaje,
        'Pedidos' : pedidos,
        'Entregados' : Entregados,
    }
   

@router.get("/ns/drivers/descargar",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_drivers(fecha_inicio : str, fecha_fin: str):
    
    results = conn.nivel_servicio_drivers(fecha_inicio,fecha_fin)
    # obtener_region
    nombre_filas = ( "Patente", "Total Rutas", "Total KM","Total Pedidos", "Total Entregados", "Total No Entregados","P EE", "Codigo1")
    nombre_excel = f"ns_driver"

    return excel.generar_excel_generico(results,nombre_filas,nombre_excel)




### Nivel Servicio Easy
@router.get("/ns/easy",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_easy():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'ns_easy')

    if datos is None:
        results = conn.detalle_pendientes_easy_hoy()
        datos = ns_easy_schema(results)
        datos_json.guardar_datos(datos,ahora,'ns_easy')

    return datos

@router.get("/ns/easy/con-ruta",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_easy_con_ruta():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'ns_easy_con_ruta')

    if datos is None:
        results = conn.detalle_pendientes_easy_hoy_con_ruta()
        datos = ns_easy_schema(results)
        datos_json.guardar_datos(datos,ahora,'ns_easy_con_ruta')

    return datos


@router.get("/ns/easy/sin-ruta",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_easy():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'ns/ns_easy_sin_ruta')

    results = conn.detalle_pendientes_easy_hoy_sin_ruta()

    datos = ns_easy_schema(results)

    return datos

@router.get("/ns/pendiente/easy/por-region",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_pendiente_easy_por_region():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'ns_easy_por_region')

    if datos is None:   

        results = conn.detalle_pendientes_easy_por_region()
        datos = ns_pendientes_easy_region_schema(results)
        datos_json.guardar_datos(datos,ahora,'ns_easy_por_region')

    return datos

@router.get("/ns/easy/panel",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_principal_ns_easy():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'panel_principal_ns_easy')

    if datos is None:   

        results = conn.panel_principal_ns_easy()
        datos = panel_principal_ns_easy(results)
        datos_json.guardar_datos(datos,ahora,'panel_principal_ns_easy')

    return datos

@router.get("/ns/easy/panel/regiones",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_regiones_ns_easy():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'panel_regiones_ns_easy')

    if datos is None:   
        results = conn.panel_regiones_ns_easy()
        datos = panel_regiones_ns_easy_schema(results)
        datos_json.guardar_datos(datos,ahora,'panel_regiones_ns_easy')
    else:
        time.sleep(9)
    return datos



@router.post("/ns/easy/descargar",status_code=status.HTTP_202_ACCEPTED)
async def descargar_ns_drivers_easy(pendientes : List[NSEasy]):
    
    tupla = [( datos_envio.Cliente, datos_envio.Guia, datos_envio.Ruta_hela, datos_envio.Direccion, 
               datos_envio.Ciudad, datos_envio.Region) for datos_envio in pendientes]

    nombre_filas = ( 'Cliente', 'Guia', "Ruta Hela","Dirección","Ciudad","Región")
    nombre_excel = f"Resumen_pendientes_en_ruta"

    return excel.generar_excel_generico(tupla,nombre_filas,nombre_excel)    


@router.get("/ns/easy/panel/no_entregados",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_no_entregados_ns_easy():
  
    results = conn.panel_no_entregados_easy()
    datos = panel_noentregas_easy_schema(results)

    return datos


### Nivel Servicio Electrolux
@router.get("/ns/electrolux",status_code=status.HTTP_202_ACCEPTED)
async def get_ns_electrolux():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'ns_electrolux')

    if datos is None:
        results = conn.detalle_pendientes_electrolux_hoy()
        datos = ns_easy_schema(results)
        datos_json.guardar_datos(datos,ahora,'ns_electrolux')

    return datos


@router.get("/ns/electrolux/panel",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_principal_ns_electrolux():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'panel_principal_ns_electrolux')

    if datos is None:   

        results = conn.panel_principal_electrolux()
        datos = panel_principal_ns_easy(results)
        datos_json.guardar_datos(datos,ahora,'panel_principal_ns_electrolux')

    return datos

@router.get("/ns/electrolux/panel/regiones",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_regiones_ns_electrolux():

    ahora = datetime.datetime.now()
    datos =  datos_json.ejecutar_por_minutos(5,'panel_regiones_ns_electrolux')

    if datos is None:   
        results = conn.panel_regiones_ns_electrolux()
        datos = panel_regiones_ns_easy_schema(results)
        datos_json.guardar_datos(datos,ahora,'panel_regiones_ns_electrolux')
    else:
        time.sleep(9)
    return datos


@router.get("/ns/electrolux/panel/no_entregados",status_code=status.HTTP_202_ACCEPTED)
async def get_panel_no_entregados_ns_electrolux():
  
    results = conn.panel_no_entregados_electrolux()
    datos = panel_noentregas_easy_schema(results)

    return datos