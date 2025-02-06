from fastapi import APIRouter, status,HTTPException, Path, UploadFile, File

from database.models.mantenedores.personal import PersonalEquipo
from database.models.mantenedores.tipo import TipoDeEquipo
from database.models.mantenedores.equipo import DescripcionEquipo
from database.models.mantenedores.asignacion import AsignarEquipo
from database.models.mantenedores.devolucion import DevolucionEquipo
from database.models.mantenedores.departamento import DepartamentoInventario
from database.models.mantenedores.sucursal import SucursalInventario
from database.models.mantenedores.estado import EstadoInventario
from database.models.mantenedores.licencia import CrearLicencia
from database.models.mantenedores.pdf import GenerarPDF
from database.models.mantenedores.asignar_entrega_acta import AsignarEntregaActa
from database.models.mantenedores.asignar_devolucion import AsignarDevolucionActa
from database.models.mantenedores.firma_acta import FirmaActa
from database.models.mantenedores.subestado import SubEstadoInventario
from database.models.mantenedores.bitacoPersona import BitacoraPersona
from database.models.mantenedores.bitacoraEquipo import BitacoraEquipo
from database.models.mantenedores.personaHabilitada import PersonaHabilitada
from database.models.mantenedores.liberar_licencia import  LiberarLicencia
from database.models.mantenedores.liberar_chip import LiberarChip
from database.models.mantenedores.liberarInsumo import LiberarInsumo

from database.schema.inventario.persona import insumos_asignado_por_serial_schema, insumo_schema, equipos_asignado_por_serial_schema, todos_los_equipo_asignado_por_persona_schema,insumo_schema,firma_entrega_schema, firma_devolucion_schema, accesorio_schema, ultima_persona_schema, crear_persona_schema, persona_equipo_schema, equipo_asignado_por_id_schema, equipo_asignado_por_persona_schema,equipo_devuelto_por_id_schema, ruta_pdf_entrega_schema, ruta_pdf_devolucion_schema
from database.schema.inventario.equipo import lista_chip_asignado_equipo_schema, lista_licencia_asignada_equipo_schema, ultimo_equipo_schema, lista_estado_schema, lista_estado_schema, lista_nr_equipo_schema,descripcion_equipo_schema, tipo_equipo_schema, lista_inventario_estado_schema, licencia_equipo_schema, folio_devolucion_schema, folio_entrega_schema, lista_subestado_schema, lista_nr_code_schema, lista_equipo_sin_join_schema, ultimo_estado_schema
from database.schema.inventario.sucursal import lista_sucursa_schema, lista_departamento_schema
##Conexiones
from database.client import reportesConnection

## en caso de generar un excel
from fastapi.responses import FileResponse
from os import getcwd, remove
import os
from os.path import dirname, join
from openpyxl import Workbook
from openpyxl.styles import Font , PatternFill, Border ,Side
from fpdf import FPDF
import json


router = APIRouter(tags=["RSV"], prefix="/api/inventario-ti")

conn = reportesConnection()

##obtener ruta de descarga de los PDF de las actas escaneadas

@router.get("/escaneado/entrega")
async def download_imprimir_planilla_entrega(id:int):
    #se realiza la busqueda por el id de asignacion y se ubica la ruta del pdf
    result = conn.read_ubicacion_pdf_entrega(id)
    if result:
    #la ruta del pdf ubicada se envia al servidor y se realiza la descarga
        return FileResponse(result[0], media_type="application/octet-stream")
    else:
        raise HTTPException(status_code=404, detail="File not found")

@router.get("/escaneado/devolucion")
async def download_imprimir_planilla_devolucion(id:int):
    #se realiza la busqueda por el id de asignacion y se ubica la ruta del pdf
    result = conn.read_ubicacion_pdf_devolucion(id)
    #la ruta del pdf ubicada se envia al servidor y se realiza la descarga
    return FileResponse(result[0], media_type="application/octet-stream")


# descarga de pdf generado por el sistema

@router.get("/descargar/entrega")
async def download_imprimir_planilla_entrega(id:int):
    #se realiza la busqueda por el id de asignacion y se ubica la ruta del pdf
    result = conn.read_ubicacion_pdf_entrega(id)
    print(result)
    return FileResponse(result[0], media_type="application/octet-stream")

@router.get("/descargar/devolucion")
async def download_imprimir_planilla_devolucion(id:int):
    #se realiza la busqueda por el id de asignacion y se ubica la ruta del pdf
    result = conn.read_ubicacion_pdf_devolucion(id)
    #la ruta del pdf ubicada se envia al servidor y se realiza la descarga
    return FileResponse(result[0][0], media_type="application/octet-stream")

##descargar excel de la planilla de personas RRHH

@router.get("/planilla/descargar")
async def download_imprimir_planilla():

    results = conn.imprimir_planilla_personas()
    wb = Workbook()
    ws = wb.active
     
    results.insert(0, ("nombres","apellidos","rut","nacionalidad", "fecha_nacimiento", "estado_civil", "telefono", "fecha_ingreso", "cargo", "domicilio", "comuna", "banco",
                       "tipo_cuenta", "numero_cuenta", "correo", "afp", "salud", "telefono_adicional","nombre_contacto", "seguro_covid", "horario", "ceco", "sueldo_base",
                       "tipo_contrato", "direccion_laboral", "enfermedad", "polera", "pantalon", "poleron", "zapato"))

    for row in results:
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save("planilla.xlsx")
    return FileResponse("planilla.xlsx")



## CARGA DE PDF SCANEADO UNA VEZ RECEPCIONADA LA FIRMA DE ENTREGA O DEVOLUCION
@router.post("/upload-entrega/{id}/{data}")
async def upload_file(id:int, data: str, file: UploadFile = File(...)):
    # se toma la ruta actual del directorio, en este caso estamos en routers
    current_directory = dirname(__file__)
    ## al estar en routers, debemos retroceder ../ y luego agregar la ruta donde se guardar los archivos
    file_path = join(current_directory, "../pdfs/acta_entrega_firma", file.filename)
    with open(file_path, "wb") as myfile:
        content = await file.read()
        conn.upload_pdf_entrega(data, id)
        myfile.write(content)
        myfile.close()
    return "success"

@router.get("/file/{name_file}")
def get_file(name_file: str):
    return FileResponse(getcwd() + "/" + name_file)


@router.get("/download/{name_file}")
def download_file(name_file: str):
    return FileResponse(getcwd() + "/" + name_file, media_type="application/octet-stream", filename=name_file)

##CREAR INFORMACION A TRAVES DE FORM

@router.post("/subEstado")
async def crear_subestado(body:SubEstadoInventario):
    try:
        data = body.dict()
        conn.ingresar_subestado(data)
        return{
             "message": "Subestado creado correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
@router.post("/asignacion")
async def asignar_equipo(body: AsignarEquipo):
    try:
        filename = os.path.basename(body.ubicacionarchivo)
        body.ubicacionarchivo = 'pdfs/foto_entrega/'+filename
        data = body.dict()
        conn.ingresar_equipo_asignado(data)
        conn.bitacora_asignar_licencia(data)
        conn.actualizar_estado_equipo(data)

        return{
             "message": "Equipo asignado correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))

## se ingresa a la tabla de asignacion los datos del chip para asi poder realizar los cambios de estado
@router.post("/asignacion-chip")
async def asignar_chip(body: AsignarEquipo):
    try:

        data = body.dict()
        print(data)
        conn.ingresar_chip_asignado(data) 
        conn.bitacora_asignar_chip(data)

        conn.actualizar_estado_chip(data)
        return{
             "message": "Equipo asignado correctamente"
        }
    except Exception as e:
        print("error")
        raise HTTPException(status_code=422, detail=str(e))
    

@router.post("/asignacion-licencia")
async def asignar_licencia(body: AsignarEquipo):
    try:

        data = body.dict() 
        conn.bitacora_asignar_licencia(data)
        conn.actualizar_estado_licencia(data)
        return{
             "message": "Licencia asignada correctamente"
        }
    except Exception as e:
        print("error")
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/asignacion-accesorio")
async def asignar_accesorio(body: AsignarEquipo):
    try:
        data = body.dict()
        conn.ingresar_accesorio_asignado(data)
        conn.bitacora_accesorios(data)
        conn.actualizar_entrega_accesorio(data)
        return{
             "message": "Accesorio asignado correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    

@router.post("/liberar-licencia")
async def liberar_licencia(body:  LiberarLicencia):
    try:
        data = body.dict()
        conn.liberar_licencia(data)
        conn.bitacora_liberar_licencia(data)
        return{
             "message": "Licencia liberada"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/liberar-chip")
async def liberar_chip(body:  LiberarChip):
    try:
        data = body.dict()
        conn.liberar_chip(data)
        conn.bitacora_liberar_chip(data)
        conn.devolver_chip(data)
        return{
             "message": "Chip liberado"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/liberar-insumo")
async def liberar_insumo(body:  LiberarInsumo):
    try:
        data = body.dict()
        conn.liberar_insumo(data)
        conn.bitacora_liberar_insumo(data)
        conn.devolver_insumo(data)
        return{
             "message": "Insumo liberado"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/licencia")
async def crear_licencia(body: CrearLicencia):
    try: 
        data=body.dict()
        conn.ingresar_licencia(data)
        return{
             "message": "Licencia creada correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))

@router.post("/personal")
async def crear_personal(body: PersonalEquipo):
    try:
        data = body.dict()
        conn.ingresar_personal_equipo(data)
        return {
            "message": "Persona agregada correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/bitacora-persona")
async def bitacora_persona(body: BitacoraPersona):
    try:
        data = body.dict()
        print(data)
        conn.bitacora_inventario_persona(data)
        return {
            "message": "Persona agregada correctamente"
        }
    except Exception as e:
        print("error", data)
        print(data)
        raise HTTPException(status_code=422, detail=str(e))
@router.post("/departamento")
async def crear_departamento(body: DepartamentoInventario):
    try: 
        data = body.dict()
        conn.ingresar_departamento(data)
        return{
            "message": "Departamento agregada correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
@router.post("/sucursal")
async def crear_sucursal(body: SucursalInventario):
    print(body)
    try:
        data = body.dict()
        print(data)
        conn.ingresar_sucursal(data)
        return{
            "message": "Sucursal agregada correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/estado")
async def crear_estado(body: EstadoInventario):
    try:
        data = body.dict()
        conn.ingresar_estado(data)
        return{
            "message": "Estado agregado correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/tipo-equipo")
async def crear_equipo(body: TipoDeEquipo ):
    try:
        data = body.dict()
        conn.ingresar_tipo_equipo(data)
        return{
            "message": "Se ha creado un equipo nuevo"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))
    
@router.post("/equipo-descripcion")
async def agregar_descripcion_de_equipo(body: DescripcionEquipo):
    try: 
        filename = os.path.basename(body.ubicacionarchivo)
        body.ubicacionarchivo = 'pdfs/foto_nuevo/'+filename
        # print(body.ubicacionarchivo )
        if body.almacenamiento == "Seleccione una opcion":
            body.almacenamiento = ""
        if body.ram == "Seleccione una opcion":
            body.ram = ""
        data = body.dict()
        print(data)
        conn.agregar_descripcion_equipo(data)
        conn.bitacora_inventario_equipo(data)
       
        return{
            "message": "Se ha añadido la descripcion del equipo"
        }
    except Exception as e :
        raise HTTPException(status_code=422, detail=str(e))
    
@router.post("/bitacora-equipo")
async def bitacora_equipo(body: BitacoraEquipo):
    try: 
        data = body.dict()

        conn.bitacora_inventario_equipo(data)
        return{
            "message": "Equipo agregado a bitacora"
        }
    except Exception as e :
        raise HTTPException(status_code=422, detail=str(e))
    
    
@router.post("/asignar-equipo")
async def asignar_equipo_personal(body: AsignarEquipo):
    try: 
        data = body.dict()
        conn.actualizar_por_entregar(data)
        conn.asignar_equipo_personal(data)
        return{
            "message": "Equipo asignado"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))
    
    
@router.post("/generar_acta_entrega")
async def crear_acta(body: AsignarEntregaActa):
    try:
        data = body.dict()
        pdf_bytes = crear_pdf(body)
        conn.update_campo_pdf_entrega(body.folio_entrega, body.id)
        return pdf_bytes 
    except Exception as e:
        print(data)
        raise HTTPException(status_code=422,detail=str(e))
    
@router.post("/generar_acta_devolucion")
async def crear_acta(body: AsignarDevolucionActa):
    try:
        data = body.dict()
        pdf_bytes = pdf_devolucion(body)
        conn.update_campo_pdf_devolucion(body.folio_devolucion, body.id)
        return pdf_bytes 
    except Exception as e:
        print(data)
        raise HTTPException(status_code=422,detail=str(e))
    


def crear_pdf(data):
    # data = json.loads(body)
    pdf = FPDF('P','mm','Letter')
    pdf.add_page()
    # pdf.set_y(-15)
    pdf.set_font('helvetica', '', 11)
    pdf.set_top_margin(15)
    pdf.set_left_margin(16)
    # pdf.set_right_margin(1,4)
    pdf.image('logo.jpg',140,20,40)
    pdf.ln(20)
    pdf.cell(80,10,'Acta de entrega', align='L')
    pdf.cell(80,10, data.folio_entrega, align='R', ln=1) 
    pdf.set_fill_color(192, 192, 192)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,9, 'DATOS DEL USUARIO', ln=1, border=True, align='C', fill=True)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Nombre', border=True , align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.nombres, border=True , align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Apellido', border=True , align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.apellidos, ln=1, border=True, align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Cargo', border=True, align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.cargo, border=True, align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'RUT', border=True, align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.rut, ln=1, border=True, align='C')
    pdf.ln(10)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,9, 'EQUIPO', border=True, ln=1 ,  align='C', fill=True)
    pdf.set_font('helvetica', 'B', 10)
    # pdf.cell(40,7,'Modelo', border=True, align='C', fill=True)
    pdf.cell(180,7,'Descripción del producto', border=True, align='C', fill=True, ln=1)
    pdf.set_font('helvetica', '', 10)
    # pdf.cell(40,7,data.marca, border=True, align='C')
    if data.info:
        for equipo in data.info:
            pdf.cell(35,7,equipo.tipo, border=True,  align='C')
            pdf.cell(35,7,equipo.marca, border=True,  align='C')
            pdf.cell(40,7,equipo.serial, border=True,  align='C')
            pdf.cell(70,7,equipo.descripcion, border=True,  align='C', ln=1)
    else:
            pdf.cell(35,7,data.marca, border=True,  align='C')
            pdf.cell(40,7,data.serial, border=True,  align='C')
            pdf.cell(105,7,data.descripcion, border=True,  align='C', ln=1)
    pdf.ln(10)
    pdf.cell(180,8,'OBSERVACIONES', border=True, ln=1, align='C', fill=True)
    pdf.cell(180,15, border=True, ln=1)
    pdf.set_font('helvetica', '', 9)
    pdf.multi_cell(180,5,""" 
                Certifico que los elementos detallados en el presente documento, me han sido entregados
                para mi cuidado y custodia con el propósito de cumplir con las tareas y asignaciones propias
                de mi cargo, siendo estos de mi única y exclusiva responsabilidad. Me comprometo a usar 
                correctamente los recursos, y solo para los fines establecidos, a no instalar ni permitir la 
                instalación de software por personal ajeno al grupo interno de trabajo. De igual forma me
                comprometo a devolver el equipo en las mismas condiciones y con los mismos accesorios 
                que me fue entregado, una vez mi vínculo laboral se dé por terminado. Está prohibido llevar 
                por cuenta propia a revisión técnica, para protección de la Garantía.
                Se prohibe el uso de número telefónico asignado fuera de Chile, si se genera consumo Roaming será cobrado
                por planilla del empleado.""",border=True)
    pdf.ln(10)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,8,'ENTREGA DE EQUIPO', border=True, ln=1, align='C', fill=True)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(90,7,'RECBIBE', border=True, align='C')
    pdf.cell(90,7,'ENTREGA', border=True, ln=1, align='C')
    pdf.set_font('helvetica', '', 10)
    pdf.cell(90,7, 'Nombre: ' +data.nombres + " " + data.apellidos, border=True)
    pdf.cell(90,7, 'Nombre :'+ data.encargado_entrega, border=True , ln=1)
    pdf.cell(90,7, 'Fecha :'+ ""+str(data.fecha_entrega), border=True)
    pdf.cell(90,7, 'Fecha :' + "" +str(data.fecha_entrega) , border=True , ln=1)
    pdf.cell(90,13, 'Firma : ', border=True)
    pdf.cell(90,13, 'Firma :', border=True , ln=1)
    # pdf.output(f"{data.folio_entrega}.pdf")
    pdf.output(f"pdfs/acta_entrega/{data.folio_entrega}.pdf")

##PDF DEVOLUCION 
def pdf_devolucion(data):
    # data = json.loads(body)
    pdf = FPDF('P','mm','Letter')
    pdf.add_page()
    # pdf.set_y(-15)
    pdf.set_font('helvetica', '', 11)
    pdf.set_top_margin(15)
    pdf.set_left_margin(16)
    # pdf.set_right_margin(1,4)
    pdf.image('logo.jpg',140,20,40)
    pdf.ln(20)
    pdf.cell(80,10,'Acta de devolución', align='L')
    pdf.cell(80,10, data.folio_devolucion, align='R', ln=1) 
    pdf.set_fill_color(192, 192, 192)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,9, 'DATOS DEL USUARIO', ln=1, border=True, align='C', fill=True)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Nombre', border=True , align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.nombres, border=True , align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Apellido', border=True , align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.apellidos, ln=1, border=True, align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'Cargo', border=True, align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.cargo, border=True, align='C')
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(35,7, 'RUT', border=True, align='C', fill=True)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(55,7,data.rut, ln=1, border=True, align='C')
    pdf.ln(10)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,9, 'DATOS DEL EQUIPO', border=True ,  align='C', fill=True, ln=1)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(40,7,'Equipo', border=True, align='C', fill=True)
    pdf.cell(40,7,'Modelo', border=True, align='C', fill=True)
    pdf.cell(40,7,'Serial', border=True, align='C', fill=True)
    pdf.cell(60,7,'Descripción', border=True, align='C', fill=True , ln=1)
    # pdf.cell(180,7,'Descripción del producto', border=True, align='C', fill=True, ln=1)
    pdf.set_font('helvetica', '', 10)
    # pdf.cell(40,7,data.marca, border=True, align='C')
    if data.info:
        for equipo in data.info:
    #     pdf.cell(180,7,equipo, border=True,  align='C', ln=1)
            pdf.cell(35,7,equipo.tipo, border=True,  align='C')
            pdf.cell(35,7,equipo.marca, border=True,  align='C')
            pdf.cell(40,7,equipo.serial, border=True,  align='C')
            pdf.cell(70,7,equipo.descripcion, border=True,  align='C', ln=1)
    else:
            pdf.cell(35,7,data.tipo, border=True,  align='C')
            pdf.cell(35,7,data.marca, border=True,  align='C')
            pdf.cell(40,7,data.serial, border=True,  align='C')
            pdf.cell(70,7,data.descripcion, border=True,  align='C', ln=1)
    pdf.ln(10)
    pdf.cell(180,8,'OBSERVACIONES', border=True, ln=1, align='C', fill=True)
    pdf.cell(180,15, border=True, ln=1)
    pdf.set_font('helvetica', '', 9)
    pdf.multi_cell(180,5,""" 
                Certifico la devolución de los elementos que me fueron entregados para el cuidado y custodia para cumplir
                   con el propósito de mi cargo""",border=True)
    pdf.ln(10)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(180,8,'ENTREGA DE EQUIPO', border=True, ln=1, align='C', fill=True)
    pdf.set_font('helvetica', 'B', 10)
    pdf.cell(90,7,'RECBIBE', border=True, align='C')
    pdf.cell(90,7,'ENTREGA', border=True, ln=1, align='C')
    pdf.set_font('helvetica', '', 10)
    pdf.cell(90,7, 'Nombre: ' +data.nombres + " " + data.apellidos, border=True)
    pdf.cell(90,7, 'Nombre :'+ data.encargado_entrega, border=True , ln=1)
    pdf.cell(90,7, 'Fecha :'+ ""+str(data.fecha_devolucion), border=True)
    pdf.cell(90,7, 'Fecha :' + "" +str(data.fecha_devolucion) , border=True , ln=1)
    pdf.cell(90,13, 'Firma : ', border=True)
    pdf.cell(90,13, 'Firma :', border=True , ln=1)
    # pdf.output(f"{data.folio_devolucion}.pdf")
    pdf.output(f"pdfs/acta_devolucion/{data.folio_devolucion}.pdf")
 
## MOSTRANDO LISTA DE LA INFORMACION ASIGNADA
@router.get("/lista-licencia")
async def lista_licencias():
    result = conn.read_licencia_windows()
    return licencia_equipo_schema(result)

@router.get("/licencias-asignadas")
async def lista_licencias_asignadas():
    result = conn.read_licencia_asignadas()
    return licencia_equipo_schema(result)

@router.get("/licencias-no-asignadas")
async def lista_licencias_sin_asignar():
    result = conn.read_licencias_no_asignadas()
    return licencia_equipo_schema(result)

@router.get("/licencias-asignadas-a-equipos")
async def lista_licencias_asignadas_a_equipos():                                 
    result = conn.read_licencias_asignadas_a_equipos()
    return lista_licencia_asignada_equipo_schema(result)

@router.get("/chip-asignados-a-equipos")
async def lista_chips_asignados_a_equipos():                                 
    result = conn.read_chips_asignados_a_equipos()
    return lista_chip_asignado_equipo_schema(result)
##modelo ocupado para obtener los datos correspondientes al realizar la devolucion del chip
@router.get("/lista-chip-devolucion")
async def lista_chips_para_devolucion():                                 
    result = conn.read_chips_asignados_para_devolucion()
    return lista_chip_asignado_equipo_schema(result)

@router.get("/chip-by-estado")
async def lista_chips_by_estado():
    result = conn.read_chip_by_estado()
    return descripcion_equipo_schema(result)

@router.get("/chip-no-asignado")
async def lista_chip_no_asignado():
    result = conn.read_chip_no_asignado()
    return descripcion_equipo_schema(result)

@router.get("/lista-personas")
async def listar_personas():
    result = conn.read_personas()
    return  crear_persona_schema(result)
@router.get("/persona-habilitada")
async def persona_habilitada():
    result = conn.read_personas_habilitadas()
    return crear_persona_schema(result)

@router.get("/lista-tipo-equipos")
async def listar_tipo_equipo():
    result = conn.read_tipo_equipo()
    return tipo_equipo_schema(result)

@router.get("/lista-tipo-con-documentacion")
async def listar_tipo_equipo_con_documentacion():
    result = conn.read_tipo_con_documentacion()
    return tipo_equipo_schema(result)

@router.get("/lista-tipo-sin-documentacion")
async def listar_tipo_equipo_sin_documentacion():
    result = conn.read_tipo_sin_documentacion()
    return tipo_equipo_schema(result)

@router.get("/ultima-persona")
async def ultima_persona_creada():
    result = conn.read_ultima_persona()
    return ultima_persona_schema(result)

@router.get("/ultimo-equipo")
async def ultima_equipo_creado():
    result = conn.read_ultimo_equipo()
    return ultimo_equipo_schema(result)

@router.get("/lista-descripcion-equipos")
async def listar_descripcion_equipo():
    result = conn.read_descripcion_equipo()
    return descripcion_equipo_schema(result)

@router.get("/lista-descripcion-por-id/{id}")
async def listar_descripcion_por_id(id: int):
    result = conn.read_descripcion_por_id(id)
    return descripcion_equipo_schema(result)

# @router.get("/lista-equipo-by-chip/{id}")
# async def listar_descripcion_por_idchip(id: int):
#     result = conn.read_descripcion_por_idchip(id)
#     return descripcion_equipo_schema(result)

@router.get("/lista-equipos")
async def lista_equipos_sin_join():
    result = conn.read_lista_equipos()
    return lista_equipo_sin_join_schema(result)

@router.get("/lista-equipos-disponibles")
async def listar_equipos_disponibles():
    result = conn.read_equipo_disponible()
    return descripcion_equipo_schema(result)

@router.get("/lista-asignacion")
async def lista_equipo_asignado_personal():
    result = conn.read_asignaciones_personal()
    return persona_equipo_schema(result)

@router.get("/lista-accesorios-asignados")
async def lista_accesorios_asignados():
    result = conn.read_accesorios_asignados()
    return accesorio_schema(result)

@router.get("/lista-insumos-asignados")
async def lista_insumos_asignados():
    result = conn.read_insumos_asignados()
    return insumo_schema(result)

@router.get("/tabla-asignados")
async def lista_asignados_sin_join():
    result = conn.read_asignados_sin_join()
    return persona_equipo_schema(result)

@router.get("/tabla-asignados-por-id/{id}")
async def lista_asignados_sin_join_por_id(id :int):
    result = conn.read_asignados_sin_join_por_id(id)
    return persona_equipo_schema(result)

@router.get("/lista-sucursales")
async def lista_de_sucursales_por_inventario_ti():
    result = conn.read_sucursales_ti()
    return lista_sucursa_schema(result)

@router.get("/lista-departamentos")
async def lista_departamento_inventario():
    result = conn.read_departamento_inventario()
    return lista_departamento_schema(result)

@router.get("/lista-estados")
async def lista_estado_inventario():
    result = conn.read_estado_inventario()
    return lista_inventario_estado_schema(result)

@router.get("/lista-estados-devolucion")
async def listado_estado_devolucion():
    result = conn.read_estados_devolucion()
    return lista_estado_schema(result)


@router.get("/lista-estado-chip")
async def listado_estado_chip():
    result = conn.read_estado_chip()
    return lista_estado_schema(result)

@router.get("/ultimo-estado")
async def ultimo_estado():
    result = conn.read_ultimo_estado()
    return ultimo_estado_schema(result)

@router.get("/lista-estado")
async def lista_estado():
    result = conn.read_estado()
    return lista_estado_schema(result)

@router.get("/lista-subestado")
async def lista_subestado():
    result = conn.read_subestado()
    return lista_subestado_schema(result)

@router.get("/lista-subestado-chip")
async def lista_subestado_chip():
    result = conn.read_subestado_chip()
    return lista_subestado_schema(result)

@router.get("/subestado-por_id/{parent_code}")
async def read_subestado_por_id(parent_code: int):
    result = conn.read_subestado_por_id(parent_code)
    return lista_subestado_schema(result)

@router.get("/estado-devolver")
async def lista_estado_devolver():
    result = conn.read_estados_devolver()
    return lista_inventario_estado_schema(result)

@router.get("/folio/{folio}")
async def equipo_asignado_por_id(folio:str ):
    result = conn.encontrar_por_folio(folio)
    return persona_equipo_schema(result)

@router.get("/folio_entrega")
async def obtener_folio_entrega():
    result = conn.read_folio_entrega()
    return folio_entrega_schema(result)

@router.get("/folio_devolucion")
async def obtener_folio_devolucion():
    result = conn.read_folio_devolucion()
    return folio_devolucion_schema(result)

@router.get("/nr_equipo/{tipo}")
async def obtener_nr_equipo(tipo: int):
    result = conn.read_nr_equipo(tipo)
    return lista_nr_equipo_schema(result)

@router.get("/code/{parent_code}")
async def obtener_nr_code(parent_code: int):
    result = conn.read_nr_code(parent_code)
    return lista_nr_code_schema(result)

@router.get("/equipos-generales")
async def obtener_lista_equipos_generales():
    result = conn.read_equipos_general()
    return  descripcion_equipo_schema(result)

@router.get("/equipos-by-tipo/{tipo}")
async def obtener_equipos_por_tipo(tipo :int):
    result = conn.read_equipos_by_tipo(tipo)
    return descripcion_equipo_schema(result)

@router.get("/equipos-by-persona-chip/{id}")
async def obtener_equipos_by_persona_chip(id:int):
    result = conn.read_equipos_by_persona_chip(id)
    return descripcion_equipo_schema(result)

@router.get("/equipos-by-persona-notebook/{id}")
async def obtener_equipos_by_persona_notebook(id:int):
    result = conn.read_equipos_by_persona_notebook(id)
    return descripcion_equipo_schema(result)

@router.get("/persona-by-departamento/{departamento}")
async def obtener_persona_por_departamento(departamento :int):
    result = conn.read_persona_by_departamento(departamento)
    return crear_persona_schema(result)

@router.get("/asignados/{id}")
async def obtener_asignados_por_id(id:int):
    result = conn.read_asignados_por_id(id)
    return equipo_asignado_por_id_schema(result)

@router.get("/devolucion/{id}")
async def obtener_devueltos_por_id(id:int):
    result = conn.read_devolucion_por_id(id)
    return equipo_devuelto_por_id_schema(result)

@router.get("/lista-equipos-por-persona/{id}")
async def obtener_lista_de_equipos_por_persona(id:str):
    result = conn.read_equipos_por_persona(id)
    return equipo_asignado_por_persona_schema(result)


@router.get("/all-equipos-por-persona/{rut}")
async def obtener_todos_los_equipos_por_persona_asignados(rut:str):
    result = conn.read_todos_los_equipos_asignados_por_persona(rut)
    return todos_los_equipo_asignado_por_persona_schema(result)

@router.get("/equipos-asignado-por-serial/{serial}")
async def obtener_equipos_asignados_por_serial(serial:str):
    result = conn.read_todos_los_equipos_asignados_por_serial(serial)
    return equipos_asignado_por_serial_schema(result)

@router.get("/insumo-asignado-por-serial/{serial}")
async def obtener_insumos_asignados_por_serial(serial:str):
    result = conn.read_todos_los_insumos_asignados_por_serial(serial)
    return insumos_asignado_por_serial_schema(result)

@router.get("/lista-equipos-por-persona-para-devolver/{id}")
async def obtener_lista_de_equipos_por_persona_por_devolver(id:str):
    result = conn.read_equipos_por_persona_por_devolver(id)
    return equipo_asignado_por_persona_schema(result)

@router.get("/lista-equipos-devueltos-por-persona/{id}")
async def obtener_lista_de_equipos_devueltos(id:str):
    result = conn.read_equipos_devueltos_por_persona(id)
    return equipo_asignado_por_persona_schema(result)

    

@router.get("/equipo-by-serial/{serial}")
async def obtener_equipo_por_serial(serial: str):
    result = conn.read_equipo_por_serial(serial)
    return descripcion_equipo_schema(result)

@router.get("/persona-by-rut/{rut}")
async def obtener_persona_by_rut(rut: str):
    result = conn.read_persona_por_rut(rut)
    return crear_persona_schema(result)

@router.get("/estado-actas/{id}")
async def obtener_estados_actas(id: int):
    result = conn.read_estados_entregas(id)
    return equipo_asignado_por_id_schema(result)

@router.get("/firma_entrega/{id}")
async def obtener_firma_entrega(id:int):
    result = conn.read_firma_entrega(id)
    return firma_entrega_schema(result)

@router.get("/firma_devolucion/{id}")
async def obtener_firma_devolucion(id:int):
    result = conn.read_firma_devolucion(id)
    return firma_devolucion_schema(result)

#AGREGANDO DEVOLUCION DE EQUIPO ASIGNADO

@router.put("/actualizar/devolucion")
async def agregar_devolucion_equipo(  body: AsignarDevolucionActa):
    try: 
        data = body.dict()
        conn.actualizar_entregado(data)
        conn.asignar_devolucion_equipo( data)
        return{
            "message": "Equipo devuelto"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))

@router.put("/actualizar/tipo")
async def editar_tipo(body: TipoDeEquipo):
    try: 
        data = body.dict()
        conn.editar_tipo_equipo( data)
        return{
            "message": "Tipo editado"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))

@router.put("/actualizar/departamento")
async def editar_departamento(body: DepartamentoInventario):
    try: 
        data = body.dict()
        conn.editar_departamento( data)
        return{
            "message": "Departamento editado"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))  

@router.put("/actualizar/licencia")
async def editar_licencia(body: CrearLicencia):
    try:
        data = body.dict()
        conn.editar_licencia(data)
        return{
            "message": "Licencia editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))
    
@router.put("/actualizar/sucursal")
async def editar_sucursal(body: SucursalInventario):
    try:
        data = body.dict()
        conn.editar_sucursal(data)
        return{
            "message": "Sucursal editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar/estado")
async def editar_estado(body: EstadoInventario):
    try:
        data = body.dict()
        conn.editar_estado(data)
        return{
            "message": "Estado editado"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar/subestado")
async def editar_estado(body: SubEstadoInventario):
    try:
        data = body.dict()
        conn.editar_subestado(data)
        return{
            "message": "SubEstado editado"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar/descripcion-equipo")
async def editar_descripcion_equipo(body: DescripcionEquipo):
    try: 
        data = body.dict()
        conn.editar_descripcion_equipo(data)
        conn.bitacora_inventario_equipo(data)
        return{
            "message":"Descripcion de equipo editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar/persona")
async def editar_persona(body: PersonalEquipo):
    try: 
        data = body.dict()
        conn.editar_persona(data)
        return{
            "message": "Persona editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar-acta")
async def editar_crear_acta(body: AsignarEntregaActa):
    try:
        data = body.dict()
        ##se actualiza el estado en la tabla equipo
        conn.actualizar_entregado(data)
        #se actualiza estados en la tabla de asignacion
        conn.datos_acta_entrega(data)
        ##bitacora al ingresar una asignacion y al realizar la entrega del acta
        conn.bitacora_inventario_asignacion(data)
        return{
            "message": "Persona editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
@router.put("/actualizar-devolucion")
async def editar_devolucion_acta(body: AsignarDevolucionActa):
    try:
        data = body.dict()
        ## al generar el acta solo se cambia el estado de la asignacion, los subestados y estados del equipo se cambian al 
        ##generar la firma del pdf por lo cual se mantiene hasta que se realice la firma
        conn.actualizar_devolucion(data)
        conn.datos_acta_devolucion(data)
            ##bitacora al ingresar una devolucion y al realizar la entrega del acta
        conn.bitacora_inventario_asignacion(data)
        return{
            "message": "Persona editada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
@router.put("/actualizar-firma-entrega")
async def editar_firma_entrega(body: FirmaActa):
    try:
        data= body.dict()
        #cambio del estado del equipo al generar firma
        conn.actualizar_firma_devolucion(data)
        #cambio del estado a true de la firma del acta
        conn.firma_acta_entrega(data)
        ##bitacora al ingresar firma de acta
        conn.bitacora_inventario_asignacion(data)
        return{
            "message": "Firma de acta realizada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    

@router.put("/actualizar-firma-devolucion")
async def editar_firma_devolucion(body: FirmaActa):
    try:
        if body.ubicacionarchivo: 
            filename = os.path.basename(body.ubicacionarchivo) 
        ## chequear si es necesario obligatoriamente enviar una foto, sino hay que realizar una validacion
        
            body.ubicacionarchivo = 'pdfs/foto_devolucion/'+filename
        
        data= body.dict()
        ##cambio de estado al equipo como entregado una vez realiza la firma del acta de entrega
        conn.actualizar_firma_devolucion(data)
        #cambio de estado y subestado al devolver el equipo
        conn.firma_acta_devolucion(data)
         ##bitacora al ingresar firma de acta
        conn.bitacora_inventario_devolucion(data)
        return{
            "message": "Firma de acta realizada"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/actualizar-habilitado")
async def editar_persona_habilitada(body: PersonaHabilitada):
    try:
        data = body.dict()
        conn.actualizar_habilitado(data)
        conn.bitacora_inventario_persona(data)
        return{
            "message": "Campo Habilitado modificadp"
        }
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    
@router.put("/devolucion-accesorio")
async def asignar_accesorio(body: AsignarEquipo):
    try:
        data = body.dict()
        conn.bitacora_accesorios(data)
        conn.actualizar_entrega_accesorio(data)
        conn.actualizar_devolucion_accesorio(data)
        return{
             "message": "Accesorio asignado correctamente"
        }
    except Exception as e:
        print("error", data)
        raise HTTPException(status_code=422, detail=str(e))