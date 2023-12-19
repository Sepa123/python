from fastapi import APIRouter, status,HTTPException, Path
from datetime import datetime
##Modelos y schemas
from typing import List
from database.schema.rsv.catalogo_producto import catalogos_productos_schema , codigos_por_color_schema
from database.models.rsv.catalogo_producto import CatalogoProducto

# from lib.excel_generico import generar_excel_generico
import lib.excel_generico as excel

from lib.codigo_nota_ventas import generar_codigo_ty_nota_venta_rsv
from database.models.rsv.datos_suma_estructura import BodySumaEstructura

from database.models.rsv.unidadEtiqueta import UnidadEtiqueta
from database.schema.rsv.etiquetas import etiquetas_productos_schema , datos_productos_etiquetas_schema


from database.models.rsv.despacho import Despacho
from database.schema.rsv.colores import colores_rsv_schema
from database.models.rsv.carga_rsv import CargaRSV

from database.models.rsv.lista_eliminar import ListaEliminar
from database.models.rsv.estructura import Estructura

from database.schema.rsv.sucursales import sucursales_rsv_schema

from database.schema.rsv.inventario_sucursal import inventarios_sucursal_schema

from database.schema.rsv.obtener_etiqueta_carga import obtener_etiquetas_carga_schema

from database.schema.rsv.cargas_rsv import cargas_rsv_schema , lista_cargas_schema

from database.schema.rsv.tipo_despacho import tipos_despacho_schema

from database.schema.rsv.datos_carga_etiqueta import datos_cargas_etiquetas_schema

from database.schema.rsv.estructura_rsv import estructuras_rsv_schema

from database.models.rsv.nota_venta import NotaVenta, NotaVentaProducto
from database.models.rsv.update_entrega_nota_venta import BodyEntregaNotaVenta

from database.schema.rsv.nota_venta import notas_ventas_schema, nota_venta
from database.schema.rsv.detalle_venta import detalle_ventas_schema
from database.schema.rsv.nota_venta_producto import nota_ventas_productos_schema

from database.models.rsv.datos_existencia_stock import ExistenciaStock

from database.schema.rsv.productoPickeado import pickeado_rsv_schema
from database.models.rsv.asignarUbicacion import RsVUbicacion
from database.schema.rsv.sucursalPorId import sucursales_rsvPorId_schema
from database.schema.rsv.verificarMatchSucursal import match_sucursales_rsv_schema

from database.schema.rsv.codigo_factura_nota_venta import generar_codigo_factura_nota_venta

from database.schema.rsv.peso_posicion_suc import peso_posicion_sucursales_schema

from database.schema.rsv.paquetesAbiertos import paquetes_abiertos_sucursal_schema
from database.models.rsv.dataAbrirPaquete import bodyPaqueteYBitacora

from database.schema.rsv.tipo_estructura import tipos_estructuras_schema
from database.schema.rsv.cantidad_actual import cant_productos_actual_schema

from database.models.rsv.armar_venta import ArmarVenta
from database.schema.rsv.armar_venta import armar_venta_schema


from database.schema.rsv.reporte_etiquetas import reporte_etiquetas_schema

from database.schema.rsv.obtener_ubicacion_cantidad import obtener_ubicacion_cantidad_schema
from database.models.rsv.obtener_ubicacion_cantidad import ObtUbicacionCantidad
##Conexiones
from database.client import reportesConnection

## en caso de generar un excel
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font , PatternFill, Border ,Side

router = APIRouter(tags=["RSV"], prefix="/api/rsv")

conn = reportesConnection()

@router.get("/catalogo")
async def obtener_catalogo_rsv():
    result = conn.read_catalogo_rsv()
    return catalogos_productos_schema(result)


## buscar catalogo por color
@router.get("/catalogo/color")
async def obtener_catalogo_rsv(color : int):
    result = conn.read_catalogo_by_color_rsv(color)
    return codigos_por_color_schema(result)


@router.get("/catalogo/color/sin_filtro")
async def obtener_catalogo_sin_filtro_rsv():
    result = conn.read_catalogo_by_color_sin_filtro_rsv()
    return codigos_por_color_schema(result)


@router.get("/colores")
async def obtener_colores_rsv():
    result = conn.read_colores_rsv()
    return colores_rsv_schema(result)


@router.get("/buscar/{codigo}")
async def buscar_producto_existente(codigo : str):
    result = conn.buscar_producto_existente_rsv(codigo)
    if result is None:
        return { "repetido": False}
    else:
        return { 
            "repetido": True,
            "message": f"El codigo {result[0]} ya existe"}


@router.post("/agregar/producto")
async def agregar_nuevo_catalogo_rsv(body : CatalogoProducto):
    body.Codigo = body.Codigo.upper()
    data = body.dict()
    conn.insert_nuevo_catalogo_rsv(data)
    return {
        "message": "Producto agregado correctamente"
    }


@router.put("/editar/producto")
async def editar_nuevo_catalogo_rsv(body : CatalogoProducto):
    body.Codigo_Original = body.Codigo
    print(body)
    data = body.dict()
    conn.update_catalogo_rsv(data)
    print(data)
    return {
        "message": "Producto editado correctamente"
    }

@router.delete("eliminar/catalogo/{catalogo}")
async def agregar_nuevo_catalogo_rsv(catalogo):
    return ""


@router.get("/cargas")
async def obtener_carga_rsv():
    result = conn.read_cargas_rsv()
    return cargas_rsv_schema(result)


@router.get("/cargas/nombre_carga/{nombre_carga}")
async def obtener_carga_rsv(nombre_carga : str):
    result = conn.read_cargas_por_nombre_carga_rsv(nombre_carga)
    return cargas_rsv_schema(result)

## mostrar datos de etiquetas por pantalla

@router.get("/datos/etiquetas/carga/{nombre_carga}")
async def obtener_datos_carga_rsv(nombre_carga : str):
    result = conn.read_datos_carga_por_nombre_rsv(nombre_carga)
    return datos_cargas_etiquetas_schema(result)


## descargar datos de etiquetas en excel

@router.get("/datos/etiquetas/carga/{nombre_carga}/descargar")
async def obtener_datos_carga_rsv(nombre_carga : str):
    results = conn.read_datos_carga_por_nombre_rsv_descarga_excel(nombre_carga)
    nombre_filas = ( "Fecha ingreso" , "Color", "Código" , "Producto" , "Paquetes" , "Unidades" , "Total Unidades" , "Cuenta paquetes" ,
                     "Paquetes Verificados" , "Cuenta Uinidades" , "Unidades verificadas")
    nombre_excel = f"{nombre_carga}"

    return excel.generar_excel_con_titulo(results,nombre_filas,nombre_excel,nombre_carga)




@router.get("/listar/cargas")
async def obtener_carga_rsv():
    result = conn.read_lista_carga_rsv()
    return lista_cargas_schema(result)

@router.get("/listar/cargas/mes")
async def obtener_carga_rsv_por_mes(mes : str):
    result = conn.read_lista_carga_rsv_por_mes(mes)
    return lista_cargas_schema(result)

@router.post("/agregar/carga")
async def insert_carga_rsv(list_body : List[CargaRSV]):
    try:
        if list_body == []:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se ha agregado ningun producto a a la carga")
        print(list_body)
        for body in list_body:
            body.Nombre_carga = body.Nombre_carga.strip()
            nombre_carga = body.Nombre_carga
            data = body.dict()
            print(data)
            conn.insert_carga_rsv(data)
        # print(data)
        return {
            "message": f"{nombre_carga} agregada correctamente"
        }
    except:
        print(" No, pase aca")
        if list_body == []:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se ha agregado ningun producto a a la carga")
        else:

            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error al ingresar la carga")

## buscar si nombre_carga existe
@router.get("/carga/buscar")
async def buscar_carga_por_nombre(nombre_carga : str):
    result = conn.buscar_cargas_rsv(nombre_carga)
    print("resultado", result)
    if result is None:
        return { "repetido": False}
    else:
        return { 
            "repetido": True,
            "message": f"La carga {result[0]} ya existe"}
    
@router.get("/etiquetas")
async def get_etiquetas_rsv():
    result = conn.read_etiquetas_rsv()
    return etiquetas_productos_schema(result)


@router.get("/etiquetas/carga/descargar")
async def download_etiquetas_carga(nombre_carga : str, codigo: str,tipo : str):

    results = conn.obtener_etiqueta_carga_rsv(nombre_carga,codigo,tipo)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("bar_code","codigo_imp","descripcion","color"))

    for row in results:
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save("excel/etiquetas_carga.xlsx")

    return FileResponse("excel/etiquetas_carga.xlsx")


@router.get("/etiquetas/carga")
async def get_etiquetas_carga(nombre_carga : str, codigo: str):
    results = conn.obtener_etiqueta_carga_rsv(nombre_carga,codigo)
    return obtener_etiquetas_carga_schema(results)



@router.get("/generar/etiquetas")
async def generar_etiquetas_por_nombre_carga(nombre_carga :str):
    results = conn.generar_etiquitas_rsv(nombre_carga)
    return {
        "alerta": results[0][0],
        "message": results[0][1]
    }

@router.get("/datos/etiquetas/productos")
async def get_datos_productos_etiquetas(nombre_carga : str):
    results = conn.read_datos_productos_etiquetas_rsv(nombre_carga)
    return datos_productos_etiquetas_schema(results)


@router.get("/sucursales")
async def get_sucursales_rsv():
    results = conn.read_sucursales_rsv()
    return sucursales_rsv_schema(results)

@router.get("/inventario/sucursales/{sucursal}")
async def get_inventario_por_sucursal(sucursal : int):
    results = conn.obtener_inventario_por_sucursal(sucursal)
    print(len(results))
    return inventarios_sucursal_schema(results)


@router.get("/inventario/sucursales/{sucursal}/descargar")
async def get_inventario_por_sucursal(sucursal : int):
    results = conn.obtener_inventario_por_sucursal_excel(sucursal)
    nombre_filas = ( "Color", "Código", "Producto", "Paquetes", "Unidades", "Total","Ubicaciones")
    nombre_excel = f"inventario-sucursal-{sucursal}"

    # for result in results:
    #     print(result[6])
    #     result[6] = ",".join(result[6])

    return excel.generar_excel_generico(results,nombre_filas,nombre_excel)
   

## eliminar cargas
@router.put("/eliminar/cargas")
async def update_carga(lista : ListaEliminar):

    if lista.lista == '':
        return {
        "message" : "no hay nada que eliminar"
    }
    codigos = lista.lista.split(',')

    print(codigos)

    print(', '.join(['%s']*len(codigos)))
    print(lista.nombre_carga)

    results = conn.delete_cargas(lista.lista, lista.nombre_carga)
    return {
        "message" : f"Cargas eliminadas ,{results}"
    }

@router.put("/editar/carga")
async def update_carga(list_body : List[CargaRSV]):
    try:
        if list_body == []:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se ha agregado ningun producto a a la carga")
        print(list_body)
        for body in list_body:
            data = body.dict()
            nombre_carga = body.Nombre_carga
            print(data)
            #verificar si producto existe en esta carga
            check = conn.check_codigo_existente_carga(nombre_carga, body.Codigo)
            print(check)
            # update de codigo existente
            if len(check) != 0:
                conn.update_carga_rsv(data)
                print("editado")
            else:
                print("insertar")
                conn.insert_carga_rsv(data)
            
        # print(data)
        return {
            "message": f"{nombre_carga} actualizada correctamente"
        }
    except:
        print(" No, pase aca")
        if list_body == []:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se ha agregado ningun producto a a la carga")
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error al ingresar la carga")

@router.get("/tipo/despacho")
async def get_tipo_despacho():
    results = conn.read_tipo_despacho_rsv()
    return tipos_despacho_schema(results)


@router.post("/agregar/nota_venta")
async def get_tipo_despacho(body : NotaVenta):
    
    id_venta = conn.get_max_id_nota_venta()
    # codigo_ty = generar_codigo_ty_nota_venta_rsv(id_venta[0])
    codigo_ty = conn.obtener_codigo_factura_venta()[2]
    body.Codigo_ty = codigo_ty
    dataVenta = body.dict()

    print(dataVenta)
    conn.insert_nota_venta_rsv(dataVenta)

    print(codigo_ty)
    for producto in body.arrays:
        producto.Id_venta = id_venta[0]
        print(producto)
        data = producto.dict()
        conn.insert_nota_venta_producto_rsv(data)

    return {
        "message": "Nota de venta agregada correctamente",
        "id_venta" : id_venta[0]
    }

@router.post("/verificar/existencia/producto")
async def verificar_existencia_producto(body : ExistenciaStock):
    data = body.dict()

    results = conn.evaluar_pedido_unidad(data)
    return {
        "Retorno" : results[0],
        "Codigo_r" : results[1],
        "Cantidad" : results[2],
        "Paquetes" : results[3],
        "Unidades" : results[4],
        "Mensaje" : results[5]
    }

@router.post("/verificar/stock/producto")
async def obtener_stock_de_producto_de_sucursal(body : ExistenciaStock):
    data = body.dict()

    results = conn.obtener_stock_de_producto_de_sucursal(body.Sucursal, body.Codigo_producto)
    return {
        "Retorno" : results[0],
        "Glosa" : results[1],
        "E_paquetes" : results[2],
        "E_unidades" : results[3],
        "Total" : results[4]
    }


@router.get("/obtener/factura/venta")
async def codigo_factura_venta():
    results = conn.obtener_codigo_factura_venta()
    return {
        "Retorno" : results[0],
        "Mensaje" : results[1],
        "Codigo" : results[2]
    }

##Y
@router.get("/producto/{barCode}")
async def obtener_producto_ById(barCode: str):
    result = conn.read_cargas_rsv_porId(barCode)
    return pickeado_rsv_schema(result)

@router.put("/producto/editar")
async def editar_ubicacion_rsv(body: RsVUbicacion):
    data = body.dict()
    conn.update_carga_rsv_porId(data)
    return {
        "message": "Producto editado correctamente"
    }

@router.get("/producto")
async def obtener_sucursal():
    result = conn.read_sucursal()
    return sucursales_rsv_schema(result)

@router.get("/sucursal/{id}")
async def obtener_sucursal_ById(id: int):
    result = conn.read_sucursal_porId(id)
    return sucursales_rsvPorId_schema(result)


@router.get("/sucursal/valida/{barCode}")
async def match_sucursal(barCode: str):
    result = conn.read_sucursal_match(barCode)
    return match_sucursales_rsv_schema(result)

##lista de paquetes abiertos

@router.get("/lista-paquetes/{sucursal}")
async def get_paquetes_abiertos(sucursal : int):
    results = conn.read_paquetes_abiertos(sucursal)
    print(results)
    return paquetes_abiertos_sucursal_schema(results)

@router.get("/etiquetas/reimprimir/descargar")
async def download_reimprimir_etiquetas_rsv(codigo: int):

    results = conn.reimprimir_etiqueta_paquete_abierto_rsv(codigo)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("bar_code","codigo_imp","descripcion","color"))
    print(results)

    for row in results:
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save("excel/etiquetas_carga.xlsx")
    return FileResponse("excel/etiquetas_carga.xlsx")

@router.get("/unica/etiqueta/descargar")
async def download_reimprimir_etiqueta_unica_rsv(nombre_carga : str, codigo: str,tipo : str, bar_code: str):

    results = conn.reimprimir_etiqueta_unica_rsv(nombre_carga, codigo, tipo, bar_code)
    wb = Workbook()
    ws = wb.active
    results.insert(0, ("bar_code","codigo_imp","descripcion","color"))
    print(results)

    for row in results:
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # get column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save("excel/etiquetas_carga.xlsx")
    return FileResponse("excel/etiquetas_carga.xlsx")



@router.get("/abrir/paquete/{bar_code}")
async def abrir_paquete_etiquetas_nuevas(bar_code: str):
    result = conn.abrir_paquete_nuevo_rsv( bar_code)
    return {
        "alerta": result[0][1],
        "message": result[0][0]
    }


@router.put("/bitacora/rsv", status_code=status.HTTP_202_ACCEPTED)
async def actualizar_bitacora_rsv(body: bodyPaqueteYBitacora):
    print(body)
    try:
        data = body.dict()
        print(data)
        conn.insert_data_bitacora_rsv(data)
    except:
          print("error en /bitacora/rsv")
          raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Error con la verificación")


@router.get("/etiqueta")
async def obtener_producto():
    result = conn.read_etiquetas_rsv()
    return etiquetas_productos_schema(result)


##/Y

@router.get("/lista/estructura")
async def get_estructura_rsv():
    results = conn.obtener_estructuras_rsv()
    return estructuras_rsv_schema(results)


@router.get("/notas_ventas/lista")
async def get_notas_ventas_by_mes_y_sucursal(mes: str, sucursal : int):
    results = conn.read_lista_ventas_por_mes(mes,sucursal)

    return notas_ventas_schema(results)



@router.get("/notas_ventas/detalle/lista/barcode")
async def get_notas_ventas_detalle_barcode(id_venta : int):
    results = conn.obtener_lista_detalle_ventas_barcode_rsv(id_venta)

    return detalle_ventas_schema(results)


@router.get("/notas_ventas/detalle/lista")
async def get_notas_ventas_detalle(id_venta : int):
    results = conn.obtener_lista_detalle_ventas_rsv(id_venta)

    return nota_ventas_productos_schema(results)


@router.get("/codigo/factura/venta")
async def get_notas_ventas_detalle(retiro : int):
    results = conn.generar_codigo_factura_nota_venta_rsv(retiro)
    return generar_codigo_factura_nota_venta(results)

@router.put("/actualizar/estado/nota_venta")
async def update_estado_entrega_nota_venta(body : BodyEntregaNotaVenta) :
    # Obtén la fecha actual
    fecha_actual = datetime.now()

    # Formatea la fecha en el formato "yyyy-mm-dd"
    fecha_formateada = fecha_actual.strftime("%Y-%m-%d")
    print(fecha_actual)
    print(fecha_formateada)

    body.Fecha_despacho = fecha_formateada

    data = body.dict()
    conn.update_estado_entrega_nota_venta_rsv(data)

    print(data)
    return {
        "message": "Venta cerrada correctamente",
        "Fecha_p" : fecha_formateada,
        "Fecha_sin_f:" : fecha_actual
    }

# peso_posicion_sucursal
@router.get("/peso/posicion/sucursal")
async def get_peso_posicion_sucursal(estructura : str, sucursal : int):
    results = conn.peso_posicion_sucursal(estructura,sucursal)

    return peso_posicion_sucursales_schema(results)


#  suma  peso_posicion_sucursal
@router.post("/peso/posicion/sucursal/suma")
async def get_peso_posicion_sucursal(body: BodySumaEstructura ):

    Suma_derecha = conn.obtener_suma_estructuras(body.Izquierda, body.Estructura, body.Sucursal)
    Suma_izquerda = conn.obtener_suma_estructuras(body.Derecha, body.Estructura, body.Sucursal)

    return {
        "Suma_izquerda" : Suma_izquerda,
        "Suma_derecha" : Suma_derecha

    }

# peso_posicion_sucursal
@router.get("/tipo/estructura")
async def get_peso_posicion_sucursal():
    results = conn.obtener_tipo_estructuras_rsv()
    return tipos_estructuras_schema(results)

# /notas_ventas/lista/completa
@router.get("/notas_ventas/lista/completa")
async def get_lista_venta_rsv(sucursal : int):
    results = conn.read_lista_ventas_rsv(sucursal)
    return notas_ventas_schema(results)

## obtener nota de venta por id

@router.get("/notas_ventas")
async def get_lista_venta_rsv(id : int):
    results = conn.obtener_nota_ventas_rsv(id)
    return nota_venta(results)

# Actualizar estructura
@router.put("/actualizar/estructura")
async def get_peso_posicion_sucursal( body  :Estructura):

    data = body.dict()
    results = conn.update_estructura_rsv(data)
    if results != 0:
        return {
            "message": "Actualizacion Realizada correctamente"
        }
    else :
        return {
            "message": "Error al actualizar"
               }       

##cambiar estado de unidades con etiquetas 

@router.put("/actualizar_unid_con_etiquetas")
async def actualizar_estado( body: UnidadEtiqueta):
    try:
        data = body.dict()
        conn.actualizar_unidad_con_etiqueta(data)
        return{
            "message": "Estado cambiado"
        }
    except Exception as e:
        raise HTTPException(status_code=422,detail=str(e))

##lista de unidades sin etiquetas

@router.get("/catalogo-unidades-sin-etiqueta")
async def obtener_unidades_sin_etiquetas_rsv():
    result = conn.read_unidades_sin_etiqueta_rsv()
    return catalogos_productos_schema(result)

@router.post("/ingresar/despacho")
async def ingresar_despacho_rsv_xd(body : Despacho):
    return ""

@router.get("/obtener/cantidad/producto/{id_venta}")
async def ingresar_despacho_rsv(id_venta : int):

    results = conn.obtener_cantidad_producto_actual_rsv(id_venta)
    return cant_productos_actual_schema(results)


@router.post("/armar/venta" )
async def armar_venta(body : ArmarVenta):
    result = conn.armar_venta_rsv(body.Nota_venta, body.Sucursal)
    return armar_venta_schema(result)


@router.get("/obtener/codigo/barcode/{bar_code}")
async def obtener_codigo_por_bar_code(bar_code : str):
    results = conn.obtener_codigo_por_bar_code(bar_code)
    codigo = ''
    if results is not None:
        codigo = results[0]

    return {
        "Codigo" : codigo
    } 



@router.post("/tomar/unidades/paquete")
async def obtener_unidades_sin_etiquetas_rsv(body : Despacho):

    tipo_code = body.Bar_code.split('@')[1].split('-')[1][0]

    if tipo_code == 'U':
        unid_x_paq = 1
        unid_con_etiqueta = True
    else :
        catalogo = conn.obtener_unidades_por_paquete(body.Codigo_producto)
        unid_x_paq = catalogo[0]
        unid_con_etiqueta = catalogo[1]

    print('barcode es:',body.Bar_code)
    

    check = conn.verificar_stock_paquete(body.Bar_code)
    check_id = check[0]
    check_stock = check[1]
    body.Cantidad = unid_x_paq
    body.Id_etiqueta = check_id

    resta = unid_x_paq - (body.Unidades- body.Uni_agregadas)
    unid_total = unid_x_paq - resta

    print("Id BAR_CODE ",check_id)
    print("Stock BAR_CODE ",check_stock)
    print(body.Uni_agregadas)
    print(unid_x_paq)
    print(body.Unidades)


    print(unid_con_etiqueta)
    
    if check_stock == False:
        return {
            "message" : f"El código {body.Bar_code} esta sin stock",
            "unid_x_paq" : 0
        }
    

    if (body.Uni_agregadas + unid_x_paq) > body.Unidades:

        if body.Uni_agregadas >= body.Unidades :
            return {
            "message" : f"Ya se agregaron las unidades necesarias al código {body.Codigo_producto}",
            "unid_x_paq" : 0
            }
        
        else:
            
            body.Cantidad = unid_total

            # si la etiqueta puede tener unidades por etiqueta
            if unid_con_etiqueta == True:
                abrir = conn.abrir_paquete_nuevo_rsv(body.Bar_code)
                etiquetas = conn.obtener_etiquetas_de_paquete(check_id)
                for num in range(unid_total) :

                    body.Bar_code = etiquetas[num][0]
                    body.Cantidad = 1
                    data = body.dict()
                    conn.insert_data_despacho_rsv(data)
                    row = conn.update_stock_etiqueta_rsv(body.Bar_code)

                # agrega la diferencia si es que el codigo es de un paquete 
                return {
                    "message" : f"Se agrego la diferencia al código {body.Codigo_producto} , que es de {unid_total} unidades",
                    "unid_x_paq" : unid_total
                } 

            # row = conn.update_stock_etiqueta_rsv(body.Bar_code)
            data = body.dict()
            conn.insert_data_despacho_rsv(data)

            # agrega la diferencia si es que el codigo es de un paquete de unidades sin etiqueta
            return {
                "message" : f"Se agrego la diferencia al código {body.Codigo_producto} , que es de {unid_total} unidades",
                "unid_x_paq" : unid_total
            }

    # if body.Unidades >= unid_x_paq and check_stock == True:
    if body.Unidades >= unid_x_paq :
        row = conn.update_stock_etiqueta_rsv(body.Bar_code)
        if tipo_code == 'U':
            data = body.dict()
            conn.insert_data_despacho_rsv(data)
            
            # si la etiqueta es de unidad
        
            return {
            "message" : "Unidad agregada correctamente",
            "unid_x_paq" : 1
            }
        else:
            body.Cantidad = 1
            data = body.dict()
            conn.insert_data_despacho_rsv(data)

        # si la etiqueta es de un paquete completo
            return {
                "message" : "Unidades agregadas correctamente",
                "unid_x_paq" : unid_x_paq
            }
    
    return {
        "message" : "f"
    }


@router.put("/cambiar/estado/venta")
async def cambiar_estado_por_id(body: BodyEntregaNotaVenta):

    fecha_actual = datetime.now()
    # Formatea la fecha en el formato "yyyy-mm-dd"
    fecha_formateada = fecha_actual.strftime("%Y-%m-%d")
    row = conn.update_preparado_nota_venta_rsv(body.Id_venta, fecha_formateada)
    # results = conn.obtener_cantidad_producto_actual_rsv(id_venta)

    return row

@router.post("/verificar/stock/productos/no_preparados")
async def obtener_stock_de_producto_de_sucursal(body : ExistenciaStock):
    data = body.dict()
    
    arr_codigos = body.Codigo_producto.split(",")

    existencia = []

    for codigo in arr_codigos:
        result = conn.obtener_stock_de_producto_de_sucursal(body.Sucursal, codigo)
        json = {
            "Codigo_producto" : codigo,
            "E_paquetes" : result[2],
            "E_unidades" : result[3],
            "Total" : result[4]
        }

        existencia.append(json)

    return existencia





@router.get("/reporte/etiquetas/{sucursal_id}")
async def obtener_catalogo_rsv(sucursal_id : int):
    result = conn.read_reporte_etiquetas_rsv(sucursal_id)
    # print(len(result))
    return reporte_etiquetas_schema(result)


@router.get("/reporte/etiquetas/{sucursal_id}/descargar/{var_r}")
async def obtener_catalogo_rsv(sucursal_id : int, var_r : str):
    results = conn.read_reporte_etiquetas_rsv_excel(sucursal_id)

    nombre_filas = ( "Carga", "Barcode", "Código", "Descripción", "Color", "Tipo etiqueta", "En stock","Ubicación")
    nombre_excel = f"Reporte-Etiquetas-{sucursal_id}"

    # for result in results:
    #     print(result[6])
    #     result[6] = ",".join(result[6])

    return excel.generar_excel_generico(results,nombre_filas,nombre_excel)
    # return reporte_etiquetas_schema(result)



@router.post("/inventario/sucursales/ubicaciones")
async def obtener_ubicacion_y_cantidad_rsv(body : ObtUbicacionCantidad):
    result = conn.obtener_ubicacion_cantidad(body.Sucursal, body.Codigo)
    # print(len(result))
    return obtener_ubicacion_cantidad_schema(result)